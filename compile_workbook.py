#!/usr/bin/env python3
"""
compile_workbook.py - Build a multi-sheet xlsx workbook from simulation results.

Reads simulation_results.json from:
  - U:/MONTECARLO/data/                      (current run)
  - U:/MONTECARLO/backups/<timestamp>/data/  (N most recent backups, default 3)

Produces a single .xlsx workbook with separate sheets for:
  - Configurations (per-run metadata)
  - Headline metrics (paper-table format)
  - Per-run full metric tables
  - Per-run ROC curves (threshold sweep arrays)
  - Cross-run analysis sheets (network-load comparison, SNR robustness, scaling)

Usage
-----
    # Default: current run + 3 most recent backups, output to data dir
    python compile_workbook.py

    # Custom locations / number of backups
    python compile_workbook.py \\
        --data-dir U:/MONTECARLO/data \\
        --backup-root U:/MONTECARLO/backups \\
        --num-backups 3 \\
        --output U:/MONTECARLO/compiled_workbook.xlsx

Author: GNACODE INC, January 2026
"""

import argparse
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# Detector & metric configuration
# ============================================================================

ALL_METHODS = ['proposed', 'lipski', 'cacfar', 'oscfar', 'cusum']
METHOD_LABELS = {
    'proposed': 'TSNFA',
    'lipski':   'Lipski',
    'cacfar':   'CA-CFAR',
    'oscfar':   'OS-CFAR',
    'cusum':    'CUSUM',
}

# All metrics extracted from each run, with display name and number format
METRICS_FULL = [
    # (json_key, display_name, excel_number_format)
    ('events_detected',              'Events Detected',           '0'),
    ('detection_rate',               'Detection Rate (%)',        '0.0'),
    ('miss_rate',                    'Miss Rate (%)',             '0.0'),
    ('fp_clusters_outside',          'FP Clusters',               '0'),
    ('event_precision',              'Event Precision (%)',       '0.0'),
    ('false_alarm_rate_clusters',    'FAR clusters /hr/node',     '0.000'),
    ('true_positives',               'TP frames',                 '0'),
    ('false_positives',              'FP frames',                 '0'),
    ('redundancy_factor',            'Redundancy (TP/event)',     '0.00'),
    ('false_alarm_rate_frames',      'FAR frames /hr/node',       '0.00'),
    ('frame_precision',              'Frame Precision (%)',       '0.000'),
    ('latency_mean_ms',              'Mean Latency (ms)',         '0.00'),
    ('latency_99th_ms',              '99th %ile Latency (ms)',    '0.00'),
    ('network_load_bytes_per_hour',  'Network Load (B/hr)',       '#,##0'),
]

# Subset of metrics to use for the "Headline" paper-table sheet
METRICS_HEADLINE = [
    'detection_rate',
    'event_precision',
    'fp_clusters_outside',
    'false_alarm_rate_clusters',
    'redundancy_factor',
    'latency_mean_ms',
    'network_load_bytes_per_hour',
]


# ============================================================================
# Styling helpers
# ============================================================================

# Common styles used across sheets
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HDR_FILL = PatternFill('solid', start_color='305496')          # dark blue
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

SUBHDR_FONT = Font(name='Arial', bold=True, size=10)
SUBHDR_FILL = PatternFill('solid', start_color='D9E1F2')       # light blue
SUBHDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

DATA_FONT = Font(name='Arial', size=10)
DATA_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
DATA_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
DATA_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

NOTE_FONT = Font(name='Arial', italic=True, size=9, color='606060')

THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)

# Highlight TSNFA row in detector tables
TSNFA_FILL = PatternFill('solid', start_color='E2EFDA')        # light green


def auto_size_columns(ws, min_width=10, max_width=40):
    """Set sensible column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            v = str(cell.value)
            # Approximate length, accounting for wrap
            if '\n' in v:
                max_len = max(max_len, max(len(line) for line in v.split('\n')))
            else:
                max_len = max(max_len, len(v))
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = width


def style_header_row(ws, row_idx, ncols, font=HDR_FONT, fill=HDR_FILL):
    """Apply header styling to a single row."""
    for col_idx in range(1, ncols + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = font
        c.fill = fill
        c.alignment = HDR_ALIGN
        c.border = THIN_BORDER
    ws.row_dimensions[row_idx].height = 28


# ============================================================================
# Loading
# ============================================================================

def load_run_json(json_path: Path) -> Optional[Dict]:
    """Load and parse one simulation_results.json."""
    if not json_path.exists():
        return None
    try:
        with open(json_path, 'r') as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        print(f"  WARNING: could not parse {json_path}: {e}")
        return None

    sim_params = data.get('_simulation_parameters', {})
    snr_db = sim_params.get('events', {}).get('snr_db')
    preset = sim_params.get('preset')
    proposed_params = sim_params.get('proposed_method', {})
    comparator_pool = sim_params.get('comparator_pool', {})

    # Find the per-network-size results block (numeric top-level key)
    num_nodes = None
    inner = None
    for k, v in data.items():
        if isinstance(k, str) and k.isdigit():
            num_nodes = int(k)
            inner = v
            break
    if num_nodes is None or inner is None:
        print(f"  WARNING: no per-network-size block in {json_path}")
        return None

    duration_hours = inner.get('config', {}).get('duration_hours')
    num_runs = inner.get('num_runs')
    total_events_obj = inner.get('true_events', {})
    total_events = total_events_obj.get('total') if isinstance(total_events_obj, dict) else None
    if isinstance(total_events, dict):
        total_events = total_events.get('mean')
    network_obj = inner.get('network', {})

    # Pull each detector's metrics
    methods = {}
    for method_key in ALL_METHODS:
        if method_key not in inner:
            continue
        m = inner[method_key]
        per_metric = {}
        for metric_key, _, _ in METRICS_FULL:
            v = m.get(metric_key)
            if v is None:
                per_metric[metric_key] = (None, None)
                continue
            if isinstance(v, dict):
                per_metric[metric_key] = (v.get('mean'), v.get('std'))
            else:
                per_metric[metric_key] = (float(v), 0.0)
        methods[method_key] = per_metric

    # Pull ROC sweep arrays (these are not aggregated; they come from first MC run only)
    roc_sweep = inner.get('roc_sweep', {})

    return {
        'num_nodes':       num_nodes,
        'snr_db':          snr_db,
        'duration_hours':  duration_hours,
        'preset':          preset,
        'num_mc_runs':     num_runs,
        'total_events':    total_events,
        'methods':         methods,
        'roc_sweep':       roc_sweep,
        'proposed_params': proposed_params,
        'comparator_pool': comparator_pool,
        'congestion_per_day': network_obj.get('congestion_per_day'),
        'source_path':     str(json_path),
        'mtime':           json_path.stat().st_mtime,
    }


def find_backup_runs(backup_root: Path, num_to_keep: int) -> List[Path]:
    """Find the N most recent backup folders containing simulation_results.json."""
    if not backup_root.exists():
        return []
    candidates = []
    for entry in backup_root.iterdir():
        if not entry.is_dir():
            continue
        json_path = entry / 'data' / 'simulation_results.json'
        if json_path.exists():
            candidates.append((json_path.stat().st_mtime, json_path))
    candidates.sort(key=lambda x: x[0], reverse=True)
    return [p for _, p in candidates[:num_to_keep]]


def collect_runs(data_dir: Path, backup_root: Path, num_backups: int) -> List[Dict]:
    """Collect (current_run + N backups) into a list of run dicts."""
    runs = []
    current_json = data_dir / 'simulation_results.json'
    if current_json.exists():
        run = load_run_json(current_json)
        if run is not None:
            run['origin'] = 'current'
            runs.append(run)
            print(f"  Loaded current run: "
                  f"{run['num_nodes']}n {run['snr_db']}dB ({run['duration_hours']}h)")
    backup_paths = find_backup_runs(backup_root, num_backups)
    for p in backup_paths:
        run = load_run_json(p)
        if run is None:
            continue
        run['origin'] = p.parent.parent.name
        runs.append(run)
        print(f"  Loaded backup '{run['origin']}': "
              f"{run['num_nodes']}n {run['snr_db']}dB ({run['duration_hours']}h)")
    return runs


def run_label(run: Dict) -> str:
    """Compact label for a run, used in column headers."""
    nn = run.get('num_nodes', '?')
    snr = run.get('snr_db', '?')
    if isinstance(snr, float):
        snr = f"{snr:.0f}"
    return f"{nn}n / {snr}dB"


def run_long_label(run: Dict) -> str:
    """Longer label including origin tag, used for individual sheets."""
    base = run_label(run)
    origin = run.get('origin', '?')
    if origin == 'current':
        return f"{base} (current)"
    short_origin = str(origin)[:13]
    return f"{base} ({short_origin})"


def run_sheet_name(run: Dict, prefix: str = '', existing_names: Optional[set] = None) -> str:
    """Build a Excel-safe sheet name (max 31 chars, no special chars).
    
    If existing_names is provided, appends a numeric suffix to avoid collisions.
    """
    nn = run.get('num_nodes', 'X')
    snr = run.get('snr_db', 'X')
    if isinstance(snr, float):
        snr = int(snr)
    base = f"{prefix}{nn}n_{snr}dB"
    name = base[:31]
    if existing_names is not None and name in existing_names:
        # Disambiguate with origin suffix or numeric suffix
        origin = str(run.get('origin', ''))
        if origin and origin != 'current' and origin != name:
            short_origin = origin[:6]
            suffix = f"_{short_origin}"
            name = (base + suffix)[:31]
            if name in existing_names:
                # Last resort: numeric suffix
                idx = 2
                while f"{base}_{idx}"[:31] in existing_names:
                    idx += 1
                name = f"{base}_{idx}"[:31]
        else:
            idx = 2
            while f"{base}_{idx}"[:31] in existing_names:
                idx += 1
            name = f"{base}_{idx}"[:31]
    if existing_names is not None:
        existing_names.add(name)
    return name


# ============================================================================
# Sheet writers
# ============================================================================

def write_readme_sheet(wb: Workbook, runs: List[Dict], extracted_at: str):
    """Sheet 1: README — overview, sheet index, run identifiers, notes."""
    ws = wb.active
    ws.title = "README"

    # Title
    ws['A1'] = "TSNFA Monte Carlo Compiled Workbook"
    ws['A1'].font = Font(name='Arial', bold=True, size=16, color='305496')
    ws.merge_cells('A1:E1')

    ws['A2'] = f"Generated: {extracted_at}"
    ws['A2'].font = NOTE_FONT
    ws.merge_cells('A2:E2')

    row = 4

    # Section: data sources
    ws.cell(row=row, column=1, value="Data Sources").font = Font(name='Arial', bold=True, size=12, color='305496')
    row += 1
    ws.cell(row=row, column=1, value=f"Number of runs compiled:")
    ws.cell(row=row, column=2, value=len(runs))
    row += 1
    for r in runs:
        ws.cell(row=row, column=1, value=f"  {run_long_label(r)}")
        ws.cell(row=row, column=2, value=r.get('source_path'))
        ws.cell(row=row, column=2).font = Font(name='Consolas', size=9)
        row += 1
    row += 1

    # Section: sheet index
    ws.cell(row=row, column=1, value="Sheet Index").font = Font(name='Arial', bold=True, size=12, color='305496')
    row += 1

    sheet_index = [
        ('README',                  'This page'),
        ('Configurations',          'Per-run metadata: SNR, network size, preset, MC runs, algorithm parameters'),
        ('Headline_Metrics',        '7 key metrics × 5 detectors × all runs (paper-table format)'),
        ('All_Metrics_<run>',       'Full 14-metric table for each run, mean ± std across MC replicates'),
        ('ROC_<run>',               'ROC threshold sweep: (threshold, DR%, FAR clusters/hr/node) per detector'),
        ('Network_Load_Comparison', 'Per-detector load × network size, with cross-detector ratios'),
        ('SNR_Robustness',          'Detection rate as function of SNR, by detector and network size'),
        ('Scaling_Analysis',        'Per-detector growth from 10n→50n at each SNR'),
    ]
    for sheet_name, description in sheet_index:
        ws.cell(row=row, column=1, value=sheet_name).font = Font(name='Consolas', bold=True, size=10)
        ws.cell(row=row, column=2, value=description)
        row += 1
    row += 1

    # Section: detector legend
    ws.cell(row=row, column=1, value="Detector Pool").font = Font(name='Arial', bold=True, size=12, color='305496')
    row += 1
    detector_info = [
        ('TSNFA',     'proposed', 'Two-stage cascade: Defence-1 (FFT-bin spectral consistency) + Defence-2 (median CFAR rank, k=N/2)'),
        ('Lipski',    'lipski',   'Lipski et al. 2021, FFT-energy detector, μ + k·σ threshold per bin (k=3 canonical), DOI 10.1109/TAES.2020.3040059'),
        ('CA-CFAR',   'cacfar',   'Finn & Johnson 1968, sliding-mean noise estimate, P_fa=1e-3 (α=7.71)'),
        ('OS-CFAR',   'oscfar',   'Rohling 1983, 75th-percentile rank-order noise estimate (k=24/32), P_fa=1e-3 (α=37.33), DOI 10.1109/TAES.1983.309350'),
        ('CUSUM',     'cusum',    'Torre 2023 Tartakovsky variant, change-point detection, α_fa=1e-5, K_end=100, DOI 10.1109/RADAR54928.2023.10371059'),
    ]
    for label, key, descr in detector_info:
        ws.cell(row=row, column=1, value=label).font = Font(name='Arial', bold=True, size=10)
        ws.cell(row=row, column=2, value=descr).font = Font(name='Arial', size=9)
        row += 1
    row += 1

    # Section: notes
    ws.cell(row=row, column=1, value="Notes").font = Font(name='Arial', bold=True, size=12, color='305496')
    row += 1
    notes = [
        "All metric values are mean ± std across Monte Carlo replicate runs (typically N=5).",
        "FAR clusters /hr/node groups consecutive false-positive triggers within 5-second window into one cluster.",
        "Event Precision = events_detected / (events_detected + FP_clusters); the headline precision metric.",
        "Network Load = per-node trigger payload bytes/hour (multiply by N-1 sensor nodes for total mesh load).",
        "ROC sweep data comes from the first MC replicate only (single-run threshold sweep).",
        "ROC curves are upper-envelope (Pareto frontier) — guaranteed monotonic.",
    ]
    for note in notes:
        ws.cell(row=row, column=1, value=f"  • {note}").font = Font(name='Arial', italic=True, size=9, color='404040')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20


def write_configurations_sheet(wb: Workbook, runs: List[Dict]):
    """Sheet: per-run config metadata."""
    ws = wb.create_sheet("Configurations")

    headers = [
        'Run', 'Origin', 'num_nodes', 'SNR (dB)', 'Duration (h)',
        'Preset', 'MC Runs', 'Total Events',
        'TSNFA gamma_d', 'TSNFA gamma_a', 'TSNFA zeta',
        'Lipski k', 'Lipski n_bins_min',
        'CA-CFAR n_ref', 'CA-CFAR n_guard', 'CA-CFAR P_fa',
        'OS-CFAR n_ref', 'OS-CFAR k_rank', 'OS-CFAR P_fa',
        'CUSUM α_fa', 'CUSUM K_end',
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for r in runs:
        cp = r.get('comparator_pool', {}) or {}
        slot1 = cp.get('slot1_lipski', {}) or {}
        slot2 = cp.get('slot2_cacfar', {}) or {}
        slot3 = cp.get('slot3_oscfar', {}) or {}
        slot4 = cp.get('slot4_cusum', {}) or {}
        pp = r.get('proposed_params', {}) or {}
        row = [
            run_long_label(r),
            r.get('origin'),
            r.get('num_nodes'),
            r.get('snr_db'),
            r.get('duration_hours'),
            r.get('preset'),
            r.get('num_mc_runs'),
            r.get('total_events'),
            pp.get('gamma_d'),
            pp.get('gamma_a'),
            pp.get('zeta'),
            slot1.get('k'),
            slot1.get('n_bins_min'),
            slot2.get('n_ref'),
            slot2.get('n_guard'),
            slot2.get('p_fa'),
            slot3.get('n_ref'),
            slot3.get('k_rank'),
            slot3.get('p_fa'),
            slot4.get('alpha_fa'),
            slot4.get('K_end'),
        ]
        ws.append(row)

    # Apply data styling
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = DATA_FONT
            c.alignment = DATA_ALIGN_CENTER if col_idx > 1 else DATA_ALIGN_LEFT
            c.border = THIN_BORDER

    auto_size_columns(ws, min_width=10, max_width=22)
    ws.freeze_panes = 'B2'


def write_headline_metrics_sheet(wb: Workbook, runs: List[Dict]):
    """Sheet: 7-row headline-metric table, one column per run."""
    ws = wb.create_sheet("Headline_Metrics")

    ws['A1'] = "Headline Metrics — TSNFA + 4 Classical Comparators"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='305496')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(runs))

    # Header row 1: run labels
    headers = ['Detector', 'Metric'] + [run_label(r) for r in runs]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4

    # Group rows by detector → metric
    for method_key in ALL_METHODS:
        for metric_idx, metric_key in enumerate(METRICS_HEADLINE):
            metric_def = next((m for m in METRICS_FULL if m[0] == metric_key), None)
            if metric_def is None:
                continue
            _, display, fmt = metric_def
            label_cell = ws.cell(row=row, column=1, value=METHOD_LABELS[method_key])
            label_cell.font = Font(name='Arial', bold=True, size=10)
            ws.cell(row=row, column=2, value=display)

            for run_idx, r in enumerate(runs, start=3):
                m = r['methods'].get(method_key, {})
                mean, std = m.get(metric_key, (None, None))
                cell = ws.cell(row=row, column=run_idx, value=mean)
                cell.number_format = fmt

            # Apply detector row styling (col 1 already has bold font set above)
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row, column=col_idx)
                if col_idx != 1:
                    c.font = DATA_FONT
                c.alignment = DATA_ALIGN_LEFT if col_idx <= 2 else DATA_ALIGN_RIGHT
                c.border = THIN_BORDER
                if method_key == 'proposed':
                    c.fill = TSNFA_FILL
            row += 1

        # Spacer row between detectors
        ws.cell(row=row, column=1, value=None)
        row += 1

    # Notes
    row += 1
    note_text = ("Notes: TSNFA rows highlighted in green. Network Load is per-node bytes/hr; "
                 "multiply by (N-1) sensor nodes for total mesh load. "
                 "Detection Rate is event-level (events with at least one trigger inside the event window).")
    nc = ws.cell(row=row, column=1, value=note_text)
    nc.font = NOTE_FONT
    nc.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + len(runs))
    ws.row_dimensions[row].height = 32

    auto_size_columns(ws, min_width=12, max_width=24)
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 26
    ws.freeze_panes = 'C4'


def write_per_run_full_metrics_sheet(wb: Workbook, run: Dict, used_names: set):
    """Sheet: full metric × detector table for a single run, mean ± std."""
    sheet_name = run_sheet_name(run, prefix='Metrics_', existing_names=used_names)
    ws = wb.create_sheet(sheet_name)

    ws['A1'] = f"All Metrics — {run_long_label(run)}"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='305496')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + 2 * len(ALL_METHODS))

    ws['A2'] = (f"Total events: {run.get('total_events') or '—'}  •  "
                f"Duration: {run.get('duration_hours')}h  •  "
                f"MC runs: {run.get('num_mc_runs')}  •  "
                f"Preset: {run.get('preset')}")
    ws['A2'].font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1 + 2 * len(ALL_METHODS))

    # Header rows 4–5: method labels (merged) + mean/std subheaders
    ws.cell(row=4, column=1, value='Metric')
    for i, method_key in enumerate(ALL_METHODS):
        col = 2 + 2 * i
        c = ws.cell(row=4, column=col, value=METHOD_LABELS[method_key])
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.cell(row=5, column=col, value='Mean')
        ws.cell(row=5, column=col + 1, value='Std')

    style_header_row(ws, 4, 1 + 2 * len(ALL_METHODS))
    for col_idx in range(1, 2 + 2 * len(ALL_METHODS)):
        c = ws.cell(row=5, column=col_idx)
        c.font = SUBHDR_FONT
        c.fill = SUBHDR_FILL
        c.alignment = SUBHDR_ALIGN
        c.border = THIN_BORDER
    ws.row_dimensions[5].height = 18

    # Body
    for metric_idx, (metric_key, display, fmt) in enumerate(METRICS_FULL):
        row = 6 + metric_idx
        ws.cell(row=row, column=1, value=display)
        for i, method_key in enumerate(ALL_METHODS):
            col = 2 + 2 * i
            mean, std = run['methods'].get(method_key, {}).get(metric_key, (None, None))
            mc = ws.cell(row=row, column=col, value=mean)
            mc.number_format = fmt
            sc = ws.cell(row=row, column=col + 1, value=std)
            sc.number_format = fmt

    # Apply row styling (highlight TSNFA columns)
    last_row = 5 + len(METRICS_FULL)
    for row_idx in range(6, last_row + 1):
        for col_idx in range(1, 2 + 2 * len(ALL_METHODS)):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = DATA_FONT
            c.alignment = DATA_ALIGN_LEFT if col_idx == 1 else DATA_ALIGN_RIGHT
            c.border = THIN_BORDER
            # TSNFA is the first method (proposed)
            if col_idx in (2, 3):
                c.fill = TSNFA_FILL

    auto_size_columns(ws, min_width=10, max_width=24)
    ws.column_dimensions['A'].width = 26
    ws.freeze_panes = 'B6'


def write_per_run_roc_sheet(wb: Workbook, run: Dict, used_names: set):
    """Sheet: ROC threshold sweep for a single run."""
    roc = run.get('roc_sweep', {})
    if not roc:
        return  # no ROC data — skip

    sheet_name = run_sheet_name(run, prefix='ROC_', existing_names=used_names)
    ws = wb.create_sheet(sheet_name)

    ws['A1'] = f"ROC Threshold Sweep — {run_long_label(run)}"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='305496')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws['A2'] = ("Each block shows one detector's threshold sweep. "
                "Curves are upper-envelope (Pareto frontier) — DR is monotonic in FAR. "
                "Canonical operating point is at threshold ≈ 1.0.")
    ws['A2'].font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    row = 4

    for method_key in ALL_METHODS:
        if method_key not in roc:
            continue
        data = roc[method_key]
        thresholds = data.get('thresholds') or []
        event_dr = data.get('event_dr') or []
        event_dr_raw = data.get('event_dr_raw') or []
        far = data.get('fp_per_hour_per_node') or []
        fp_clusters = data.get('fp_clusters_total') or []

        # Detector header
        c = ws.cell(row=row, column=1, value=METHOD_LABELS[method_key])
        c.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        c.fill = HDR_FILL
        c.alignment = DATA_ALIGN_LEFT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 22
        row += 1

        # Sub-header
        sub_headers = ['Threshold', 'FAR clusters/hr/node', 'DR (envelope) %', 'DR (raw) %', 'FP clusters total']
        for col_idx, h in enumerate(sub_headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = SUBHDR_FONT
            c.fill = SUBHDR_FILL
            c.alignment = SUBHDR_ALIGN
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1

        # Data rows — sorted by FAR ascending (this is how the simulator stored them)
        n = len(thresholds)
        for i in range(n):
            ws.cell(row=row, column=1, value=thresholds[i] if i < len(thresholds) else None).number_format = '0.0000'
            ws.cell(row=row, column=2, value=far[i] if i < len(far) else None).number_format = '0.0000'
            ws.cell(row=row, column=3, value=event_dr[i] if i < len(event_dr) else None).number_format = '0.00'
            if i < len(event_dr_raw):
                ws.cell(row=row, column=4, value=event_dr_raw[i]).number_format = '0.00'
            ws.cell(row=row, column=5, value=fp_clusters[i] if i < len(fp_clusters) else None).number_format = '0'

            for col_idx in range(1, 6):
                c = ws.cell(row=row, column=col_idx)
                c.font = DATA_FONT
                c.alignment = DATA_ALIGN_RIGHT
                c.border = THIN_BORDER

            row += 1

        row += 1  # spacer between detector blocks

    auto_size_columns(ws, min_width=12, max_width=22)


def write_network_load_comparison_sheet(wb: Workbook, runs: List[Dict]):
    """Sheet: cross-detector network load comparison with ratios."""
    ws = wb.create_sheet("Network_Load_Comparison")

    ws['A1'] = "Network Load Comparison Across Runs"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='305496')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(runs))

    ws['A2'] = ("Per-node load (B/hr), total mesh load (= per-node × (N-1) sensor nodes), "
                "and ratio vs TSNFA. Cells use formulas referencing run data above.")
    ws['A2'].font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4 + len(runs))

    # Section 1: per-node load
    row = 4
    ws.cell(row=row, column=1, value="Per-node Network Load (B/hr)").font = Font(name='Arial', bold=True, size=11, color='305496')
    row += 1
    headers = ['Detector'] + [run_label(r) for r in runs]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    pernode_start_row = row
    for method_key in ALL_METHODS:
        ws.cell(row=row, column=1, value=METHOD_LABELS[method_key]).font = Font(name='Arial', bold=True, size=10)
        for run_idx, r in enumerate(runs, start=2):
            mean, _ = r['methods'].get(method_key, {}).get('network_load_bytes_per_hour', (None, None))
            c = ws.cell(row=row, column=run_idx, value=mean)
            c.number_format = '#,##0'
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col_idx)
            if not c.font.bold:
                c.font = DATA_FONT
            c.alignment = DATA_ALIGN_LEFT if col_idx == 1 else DATA_ALIGN_RIGHT
            c.border = THIN_BORDER
            if method_key == 'proposed':
                c.fill = TSNFA_FILL
        row += 1

    # Section 2: total mesh load (per-node × (N-1))
    row += 1
    ws.cell(row=row, column=1, value="Total Mesh Load (kB/hr) = per-node × (N-1)").font = Font(name='Arial', bold=True, size=11, color='305496')
    row += 1
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    total_start_row = row
    for method_idx, method_key in enumerate(ALL_METHODS):
        ws.cell(row=row, column=1, value=METHOD_LABELS[method_key]).font = Font(name='Arial', bold=True, size=10)
        # Formula: per-node value × (N-1) / 1000 to convert to kB
        for run_idx, r in enumerate(runs, start=2):
            n_nodes = r.get('num_nodes', 0) or 0
            sensor_nodes = max(0, n_nodes - 1)
            pernode_cell = ws.cell(row=pernode_start_row + method_idx, column=run_idx).coordinate
            formula = f"={pernode_cell}*{sensor_nodes}/1000"
            c = ws.cell(row=row, column=run_idx, value=formula)
            c.number_format = '#,##0.0'
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col_idx)
            if not c.font.bold:
                c.font = DATA_FONT
            c.alignment = DATA_ALIGN_LEFT if col_idx == 1 else DATA_ALIGN_RIGHT
            c.border = THIN_BORDER
            if method_key == 'proposed':
                c.fill = TSNFA_FILL
        row += 1

    # Section 3: ratio vs TSNFA (per-node)
    row += 1
    ws.cell(row=row, column=1, value="Per-node Load Ratio vs TSNFA").font = Font(name='Arial', bold=True, size=11, color='305496')
    row += 1
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    for method_idx, method_key in enumerate(ALL_METHODS):
        ws.cell(row=row, column=1, value=METHOD_LABELS[method_key]).font = Font(name='Arial', bold=True, size=10)
        for run_idx, r in enumerate(runs, start=2):
            method_cell = ws.cell(row=pernode_start_row + method_idx, column=run_idx).coordinate
            tsnfa_cell = ws.cell(row=pernode_start_row + 0, column=run_idx).coordinate  # TSNFA is index 0
            formula = f"=IFERROR({method_cell}/{tsnfa_cell},\"-\")"
            c = ws.cell(row=row, column=run_idx, value=formula)
            c.number_format = '0.0"×"'
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col_idx)
            if not c.font.bold:
                c.font = DATA_FONT
            c.alignment = DATA_ALIGN_LEFT if col_idx == 1 else DATA_ALIGN_RIGHT
            c.border = THIN_BORDER
            if method_key == 'proposed':
                c.fill = TSNFA_FILL
        row += 1

    # Note
    row += 1
    note_text = ("Per-node load is the bytes per hour each sensor node transmits (across all triggers). "
                 "Total Mesh Load is the aggregate traffic on the network, computed as per-node × (N-1) "
                 "where N-1 is the number of sensor nodes (one node is the sink). "
                 "Ratio shows how many times more bandwidth a detector consumes than TSNFA at the same operating point.")
    nc = ws.cell(row=row, column=1, value=note_text)
    nc.font = NOTE_FONT
    nc.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + len(runs))
    ws.row_dimensions[row].height = 48

    auto_size_columns(ws, min_width=12, max_width=22)
    ws.column_dimensions['A'].width = 14
    ws.freeze_panes = 'B5'


def write_snr_robustness_sheet(wb: Workbook, runs: List[Dict]):
    """Sheet: detection rate vs SNR by detector and network size."""
    ws = wb.create_sheet("SNR_Robustness")

    ws['A1'] = "SNR Robustness — Detection Rate vs SNR"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='305496')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws['A2'] = ("Each detector's detection rate (%) at the canonical operating point, "
                "across SNR values. Drops at lower SNR indicate per-node algorithmic brittleness.")
    ws['A2'].font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Determine unique sizes and SNRs
    sizes = sorted({r['num_nodes'] for r in runs if r.get('num_nodes') is not None})
    snrs = sorted({r['snr_db'] for r in runs if r.get('snr_db') is not None}, reverse=True)

    row = 4
    for size_idx, n_nodes in enumerate(sizes):
        # Section header per size
        ws.cell(row=row, column=1, value=f"Network Size: {n_nodes} nodes").font = Font(name='Arial', bold=True, size=11, color='305496')
        row += 1

        # Header row
        headers = ['Detector'] + [f"{int(s)} dB" for s in snrs] + ['SNR Drop (Δ%)']
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        for method_key in ALL_METHODS:
            ws.cell(row=row, column=1, value=METHOD_LABELS[method_key]).font = Font(name='Arial', bold=True, size=10)
            dr_values = []
            for snr_idx, snr in enumerate(snrs, start=2):
                # Find the run matching (n_nodes, snr)
                matching = [r for r in runs if r['num_nodes'] == n_nodes and r['snr_db'] == snr]
                if not matching:
                    continue
                r = matching[0]
                mean, std = r['methods'].get(method_key, {}).get('detection_rate', (None, None))
                dr_values.append(mean)
                c = ws.cell(row=row, column=snr_idx, value=mean)
                c.number_format = '0.0"%"'

            # SNR drop column (highest SNR DR - lowest SNR DR)
            if len(dr_values) >= 2 and all(v is not None for v in dr_values):
                drop = dr_values[0] - dr_values[-1]   # snrs is sorted descending, so [0]=high, [-1]=low
                drop_cell = ws.cell(row=row, column=2 + len(snrs), value=drop)
                drop_cell.number_format = '0.0"%"'
                # Highlight large drops
                if abs(drop) > 30:
                    drop_cell.fill = PatternFill('solid', start_color='FFC7CE')  # red-ish
                elif abs(drop) > 10:
                    drop_cell.fill = PatternFill('solid', start_color='FFEB9C')  # amber

            # Style cells
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row, column=col_idx)
                if not c.font.bold:
                    c.font = DATA_FONT
                c.alignment = DATA_ALIGN_LEFT if col_idx == 1 else DATA_ALIGN_RIGHT
                c.border = THIN_BORDER
                if method_key == 'proposed' and not c.fill.start_color.rgb in ('00FFC7CE', '00FFEB9C'):
                    c.fill = TSNFA_FILL
            row += 1

        row += 1  # spacer between sizes

    # Note
    note_text = ("SNR Drop column = DR(highest SNR) − DR(lowest SNR). "
                 "Cells highlighted red if drop > 30 percentage points (severe brittleness), "
                 "amber if 10–30 (moderate). TSNFA's drop should be near 0 across both sizes.")
    nc = ws.cell(row=row, column=1, value=note_text)
    nc.font = NOTE_FONT
    nc.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 36

    auto_size_columns(ws, min_width=12, max_width=20)


def write_scaling_analysis_sheet(wb: Workbook, runs: List[Dict]):
    """Sheet: per-detector growth from smallest→largest network at each SNR."""
    ws = wb.create_sheet("Scaling_Analysis")

    ws['A1'] = "Network Scaling — From Smallest to Largest Network"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='305496')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws['A2'] = ("Per-metric growth ratio when going from smallest to largest network size at each SNR. "
                "Detection-quality metrics (DR, precision) should be invariant; only counts (FP clusters, "
                "network load) should scale linearly with N.")
    ws['A2'].font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    sizes = sorted({r['num_nodes'] for r in runs if r.get('num_nodes') is not None})
    snrs = sorted({r['snr_db'] for r in runs if r.get('snr_db') is not None}, reverse=True)

    if len(sizes) < 2:
        ws.cell(row=4, column=1, value="Need at least 2 different network sizes for scaling analysis.")
        return

    n_small, n_large = sizes[0], sizes[-1]
    metrics_to_scale = [
        ('detection_rate',              'Detection Rate (%)',           '0.0'),
        ('event_precision',             'Event Precision (%)',          '0.0'),
        ('fp_clusters_outside',         'FP Clusters',                  '0'),
        ('false_alarm_rate_clusters',   'FAR clusters /hr/node',        '0.00'),
        ('network_load_bytes_per_hour', 'Network Load (B/hr)',          '#,##0'),
        ('latency_mean_ms',             'Mean Latency (ms)',            '0.0'),
    ]

    row = 4
    for snr in snrs:
        ws.cell(row=row, column=1, value=f"SNR = {int(snr)} dB").font = Font(name='Arial', bold=True, size=11, color='305496')
        row += 1

        headers = ['Detector', 'Metric', f'{n_small}n', f'{n_large}n', 'Ratio (×)', 'Δ Absolute']
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        for method_key in ALL_METHODS:
            for metric_key, display, fmt in metrics_to_scale:
                small_run = next((r for r in runs if r['num_nodes'] == n_small and r['snr_db'] == snr), None)
                large_run = next((r for r in runs if r['num_nodes'] == n_large and r['snr_db'] == snr), None)
                if small_run is None or large_run is None:
                    continue

                v_small = small_run['methods'].get(method_key, {}).get(metric_key, (None, None))[0]
                v_large = large_run['methods'].get(method_key, {}).get(metric_key, (None, None))[0]

                ws.cell(row=row, column=1, value=METHOD_LABELS[method_key]).font = Font(name='Arial', bold=True, size=10)
                ws.cell(row=row, column=2, value=display)
                ws.cell(row=row, column=3, value=v_small).number_format = fmt
                ws.cell(row=row, column=4, value=v_large).number_format = fmt

                # Ratio formula (cell-referenced so users can see the calculation)
                small_cell = ws.cell(row=row, column=3).coordinate
                large_cell = ws.cell(row=row, column=4).coordinate
                ratio_formula = f'=IF(AND(ISNUMBER({small_cell}),ISNUMBER({large_cell}),{small_cell}<>0),{large_cell}/{small_cell},"-")'
                rc = ws.cell(row=row, column=5, value=ratio_formula)
                rc.number_format = '0.00"×"'

                delta_formula = f'=IF(AND(ISNUMBER({small_cell}),ISNUMBER({large_cell})),{large_cell}-{small_cell},"-")'
                dc = ws.cell(row=row, column=6, value=delta_formula)
                dc.number_format = fmt

                # Style cells
                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=row, column=col_idx)
                    if not c.font.bold:
                        c.font = DATA_FONT
                    c.alignment = DATA_ALIGN_LEFT if col_idx <= 2 else DATA_ALIGN_RIGHT
                    c.border = THIN_BORDER
                    if method_key == 'proposed':
                        c.fill = TSNFA_FILL
                row += 1

            row += 1  # spacer between detectors

        row += 1  # spacer between SNRs

    # Note
    note_text = (f"Ratio = {n_large}n value / {n_small}n value. Δ Absolute = difference. "
                 f"For detection-quality metrics (DR, Event Precision) the ratio should be ~1.0 "
                 f"(quality is per-node, not per-network). For count-based metrics (FP Clusters, "
                 f"Network Load), expect ratio ≈ ({n_large}-1)/({n_small}-1) = "
                 f"{(n_large - 1) / max(1, n_small - 1):.1f}× from network size alone.")
    nc = ws.cell(row=row, column=1, value=note_text)
    nc.font = NOTE_FONT
    nc.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 48

    auto_size_columns(ws, min_width=10, max_width=22)
    ws.column_dimensions['B'].width = 26


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Compile a multi-sheet xlsx workbook from simulation runs',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument('--data-dir', default='U:/MONTECARLO/data',
                        help='Directory containing the current simulation_results.json')
    parser.add_argument('--backup-root', default='U:/MONTECARLO/backups',
                        help='Root containing timestamped backup folders')
    parser.add_argument('--num-backups', type=int, default=3,
                        help='Number of most-recent backups to include')
    parser.add_argument('--output', default=None,
                        help='Path to output xlsx (defaults to <data-dir>/compiled_workbook.xlsx)')

    args = parser.parse_args()

    data_dir = Path(args.data_dir).resolve()
    backup_root = Path(args.backup_root).resolve()
    output_path = (Path(args.output).resolve() if args.output
                   else data_dir / 'compiled_workbook.xlsx')

    print('=' * 75)
    print('Compile Workbook — TSNFA Monte Carlo Results')
    print('=' * 75)
    print(f"  Data dir:      {data_dir}")
    print(f"  Backup root:   {backup_root}")
    print(f"  Backups kept:  {args.num_backups} most recent")
    print(f"  Output:        {output_path}")
    print()

    print("Loading runs...")
    runs = collect_runs(data_dir, backup_root, args.num_backups)
    if not runs:
        print("\nNo runs found. Check --data-dir and --backup-root.")
        return 1

    print(f"\nBuilding workbook with {len(runs)} runs...")
    wb = Workbook()
    extracted_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Sheet 1: README
    write_readme_sheet(wb, runs, extracted_at)
    print("  README")

    # Sheet 2: Configurations
    write_configurations_sheet(wb, runs)
    print("  Configurations")

    # Sheet 3: Headline metrics
    write_headline_metrics_sheet(wb, runs)
    print("  Headline_Metrics")

    # Per-run full metrics sheets
    used_metric_names = set()
    for r in runs:
        write_per_run_full_metrics_sheet(wb, r, used_metric_names)
        print(f"  Metrics_{r['num_nodes']}n_{int(r['snr_db'])}dB")

    # Per-run ROC sheets
    used_roc_names = set()
    for r in runs:
        if r.get('roc_sweep'):
            write_per_run_roc_sheet(wb, r, used_roc_names)
            print(f"  ROC_{r['num_nodes']}n_{int(r['snr_db'])}dB")

    # Cross-run analysis sheets
    write_network_load_comparison_sheet(wb, runs)
    print("  Network_Load_Comparison")

    write_snr_robustness_sheet(wb, runs)
    print("  SNR_Robustness")

    write_scaling_analysis_sheet(wb, runs)
    print("  Scaling_Analysis")

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    file_size_kb = output_path.stat().st_size / 1024
    print(f"\n  Workbook → {output_path} ({file_size_kb:.1f} kB)")
    print('\nDone.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())