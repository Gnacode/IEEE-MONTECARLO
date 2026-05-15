#!/usr/bin/env python3
"""
make_data_page.py - Build interactive Plotly HTML data page from the compiled workbook.

Reads compiled_workbook.xlsx (the canonical clean dataset, produced by
compile_workbook.py) and produces a self-contained data page suitable for
hosting on GitHub Pages alongside paper supplementary materials.

This is the second stage of the two-stage pipeline:
    1. compile_workbook.py:  JSONs        -> compiled_workbook.xlsx
    2. make_data_page.py:    xlsx         -> datapage/index.html + figures

Outputs (all in --output-dir):
  index.html                              Landing page with figure index, metadata, notes
  fig_01_headline_matrix.html             Overview heatmap of detection rate per config
  fig_02_roc_envelope.html                4-panel ROC curves
  fig_03_network_load.html                Bar chart of per-detector network load
  fig_04_snr_robustness.html              Paired line plot DR vs SNR
  fig_05_scaling.html                     Multi-metric scaling 10n -> 50n
  fig_06_bandwidth_precision.html         Scatter: bandwidth vs precision per detector
  fig_07_3d_quality.html                  3D scatter: FAR x Load x DR per detector
  data.json                               Compiled run data, for reference

Each figure is a single self-contained .html file with embedded Plotly JS - no CDN, no
external dependencies. Open directly in any browser.

Usage
-----
    # Default location
    python make_data_page.py

    # Custom workbook path or output dir
    python make_data_page.py \\
        --workbook U:/MONTECARLO/data/compiled_workbook.xlsx \\
        --output-dir U:/MONTECARLO/data/datapage

Requirements: plotly>=5.0, openpyxl>=3.0
    pip install plotly openpyxl

Author: GNACODE INC, January 2026
"""

import argparse
import json
import math
import re
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
except ImportError:
    print("ERROR: plotly is not installed. Install with:")
    print("    pip install plotly")
    raise SystemExit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed. Install with:")
    print("    pip install openpyxl")
    raise SystemExit(1)


# ============================================================================
# Detector & visual configuration
# ============================================================================

# Order matters: TSNFA first
ALL_METHODS = ['proposed', 'lipski', 'cacfar', 'oscfar', 'cusum']
METHOD_LABELS = {
    'proposed': 'TSNFA',
    'lipski':   'Lipski FFT',
    'cacfar':   'CA-CFAR',
    'oscfar':   'OS-CFAR',
    'cusum':    'CUSUM',
}
# Map workbook display names -> internal method keys
WORKBOOK_LABEL_TO_METHOD = {
    'TSNFA':      'proposed',
    'Lipski':     'lipski',
    'Lipski FFT': 'lipski',
    'CA-CFAR':    'cacfar',
    'OS-CFAR':    'oscfar',
    'CUSUM':      'cusum',
}
COLORS = {
    'proposed': '#2ecc71',  # green
    'lipski':   '#3498db',  # blue
    'cacfar':   '#e74c3c',  # red
    'oscfar':   '#f39c12',  # orange
    'cusum':    '#9b59b6',  # purple
}
SYMBOLS = {
    'proposed': 'circle',
    'lipski':   'square',
    'cacfar':   'triangle-up',
    'oscfar':   'diamond',
    'cusum':    'triangle-down',
}

# Workbook metric display names -> internal keys (must match what's in compile_workbook.py)
WORKBOOK_METRIC_KEYS = {
    'Events Detected':           'events_detected',
    'Detection Rate (%)':        'detection_rate',
    'Miss Rate (%)':             'miss_rate',
    'FP Clusters':               'fp_clusters_outside',
    'Event Precision (%)':       'event_precision',
    'FAR clusters /hr/node':     'false_alarm_rate_clusters',
    'TP frames':                 'true_positives',
    'FP frames':                 'false_positives',
    'Redundancy (TP/event)':     'redundancy_factor',
    'FAR frames /hr/node':       'false_alarm_rate_frames',
    'Frame Precision (%)':       'frame_precision',
    'Mean Latency (ms)':         'latency_mean_ms',
    '99th %ile Latency (ms)':    'latency_99th_ms',
    'Network Load (B/hr)':       'network_load_bytes_per_hour',
}

PLOTLY_TEMPLATE = 'plotly_white'

FIG_LAYOUT = dict(
    template=PLOTLY_TEMPLATE,
    font=dict(family='Arial, sans-serif', size=12, color='#2c2c2c'),
    title_font=dict(size=16, family='Arial, sans-serif', color='#1f3864'),
    margin=dict(l=70, r=30, t=80, b=60),
    hoverlabel=dict(bgcolor='white', font_size=12, font_family='Arial'),
    legend=dict(
        bgcolor='rgba(255,255,255,0.85)',
        bordercolor='#cccccc',
        borderwidth=1,
        font=dict(size=11),
    ),
)


# ============================================================================
# Workbook reading
# ============================================================================

# Match a sheet-name pattern produced by compile_workbook.py for run-specific
# sheets, e.g. "Metrics_50n_12dB" or "ROC_10n_18dB"
_SHEET_RUN_RE = re.compile(r'^(?P<prefix>Metrics|ROC)_(?P<n>\d+)n_(?P<snr>\d+)dB(?:_.*)?$')

# Match the column label format used in Headline_Metrics, e.g. "50n / 12dB"
_HEADLINE_LABEL_RE = re.compile(r'^\s*(?P<n>\d+)n\s*/\s*(?P<snr>\d+)dB\s*$')


def _to_int_safe(v):
    """Best-effort int conversion, returns None if not numeric."""
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _to_float_safe(v):
    """Best-effort float conversion, returns None if not numeric."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_configurations(wb) -> List[Dict]:
    """Parse the Configurations sheet -> list of run dicts.

    Each run has the metadata fields needed elsewhere: num_nodes, snr_db,
    duration_hours, preset, num_mc_runs, total_events, origin, comparator
    parameters.
    """
    if 'Configurations' not in wb.sheetnames:
        raise ValueError("Workbook missing required sheet 'Configurations'")

    ws = wb['Configurations']
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    runs = []
    for row in range(2, ws.max_row + 1):
        values = {}
        for col_idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            values[header] = ws.cell(row=row, column=col_idx).value

        # Skip empty rows
        if not values.get('Run'):
            continue

        runs.append({
            'run_label':       values.get('Run'),
            'origin':          values.get('Origin'),
            'num_nodes':       _to_int_safe(values.get('num_nodes')),
            'snr_db':          _to_float_safe(values.get('SNR (dB)')),
            'duration_hours':  _to_float_safe(values.get('Duration (h)')),
            'preset':          values.get('Preset'),
            'num_mc_runs':     _to_int_safe(values.get('MC Runs')),
            'total_events':    _to_int_safe(values.get('Total Events')),
            # Algorithm parameters (kept for index.html, not used in figures)
            'tsnfa_gamma_d':   values.get('TSNFA gamma_d'),
            'tsnfa_gamma_a':   values.get('TSNFA gamma_a'),
            'tsnfa_zeta':      values.get('TSNFA zeta'),
            'lipski_k':        values.get('Lipski k'),
            'cacfar_p_fa':     values.get('CA-CFAR P_fa'),
            'oscfar_k_rank':   values.get('OS-CFAR k_rank'),
            'oscfar_p_fa':     values.get('OS-CFAR P_fa'),
            'cusum_alpha_fa':  values.get('CUSUM α_fa'),
        })

    return runs


def read_per_run_metrics(wb, run: Dict) -> Dict[str, Dict[str, Tuple[Optional[float], Optional[float]]]]:
    """Parse a per-run Metrics_<n>n_<snr>dB sheet for one run.

    Returns: {method_key: {metric_key: (mean, std)}}
    """
    n_nodes = run['num_nodes']
    snr = run['snr_db']
    snr_int = int(snr) if snr is not None else 'X'
    target_sheet = f"Metrics_{n_nodes}n_{snr_int}dB"

    # Find the sheet, falling back to any sheet matching this run if exact name not found
    sheet_name = None
    if target_sheet in wb.sheetnames:
        sheet_name = target_sheet
    else:
        # Look for disambiguated name like Metrics_50n_12dB_xxxxx
        for s in wb.sheetnames:
            if s.startswith(target_sheet):
                sheet_name = s
                break

    if sheet_name is None:
        return {m: {} for m in ALL_METHODS}

    ws = wb[sheet_name]

    # Row 4: method labels (with Mean/Std merged across two cols each)
    # Row 5: alternating Mean/Std subheaders
    # Row 6+: one row per metric, value in (Mean col, Std col) for each method
    method_to_cols = {}  # method_key -> (mean_col, std_col)
    for col in range(2, ws.max_column + 1):
        label = ws.cell(row=4, column=col).value
        if label and label in WORKBOOK_LABEL_TO_METHOD:
            method_key = WORKBOOK_LABEL_TO_METHOD[label]
            method_to_cols[method_key] = (col, col + 1)

    methods_data = {m: {} for m in ALL_METHODS}
    for row in range(6, ws.max_row + 1):
        metric_display = ws.cell(row=row, column=1).value
        if not metric_display:
            continue
        metric_key = WORKBOOK_METRIC_KEYS.get(metric_display)
        if metric_key is None:
            continue
        for method_key, (mean_col, std_col) in method_to_cols.items():
            mean = _to_float_safe(ws.cell(row=row, column=mean_col).value)
            std = _to_float_safe(ws.cell(row=row, column=std_col).value)
            methods_data[method_key][metric_key] = (mean, std)

    return methods_data


def read_per_run_roc(wb, run: Dict) -> Dict[str, Dict]:
    """Parse a per-run ROC_<n>n_<snr>dB sheet.

    Returns: {method_key: {'thresholds':[...], 'fp_per_hour_per_node':[...],
                          'event_dr':[...], 'event_dr_raw':[...],
                          'fp_clusters_total':[...]}}
    """
    n_nodes = run['num_nodes']
    snr = run['snr_db']
    snr_int = int(snr) if snr is not None else 'X'
    target_sheet = f"ROC_{n_nodes}n_{snr_int}dB"

    sheet_name = None
    if target_sheet in wb.sheetnames:
        sheet_name = target_sheet
    else:
        for s in wb.sheetnames:
            if s.startswith(target_sheet):
                sheet_name = s
                break

    if sheet_name is None:
        return {}

    ws = wb[sheet_name]
    roc = {}

    # Walk the sheet looking for detector blocks.
    # Detector header rows: just a label in col 1 (e.g. "TSNFA"), nothing else.
    # Sub-header row immediately follows: ['Threshold', 'FAR clusters/hr/node', 'DR (envelope) %', 'DR (raw) %', 'FP clusters total']
    # Then data rows until a blank or the next detector header.
    current_method = None
    rows = list(ws.iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        row = rows[i]
        if row and row[0] in WORKBOOK_LABEL_TO_METHOD and (row[1] is None or not isinstance(row[1], str) or row[1] == ''):
            # New detector block
            current_method = WORKBOOK_LABEL_TO_METHOD[row[0]]
            # Skip the next row (subheader)
            roc[current_method] = {
                'thresholds': [],
                'fp_per_hour_per_node': [],
                'event_dr': [],
                'event_dr_raw': [],
                'fp_clusters_total': [],
            }
            i += 2
            continue

        # Data row inside a detector block
        if current_method and row and row[0] is not None:
            thr = _to_float_safe(row[0])
            far = _to_float_safe(row[1] if len(row) > 1 else None)
            dr_env = _to_float_safe(row[2] if len(row) > 2 else None)
            dr_raw = _to_float_safe(row[3] if len(row) > 3 else None)
            fp_clusters = _to_float_safe(row[4] if len(row) > 4 else None)
            if thr is not None and dr_env is not None:
                roc[current_method]['thresholds'].append(thr)
                roc[current_method]['fp_per_hour_per_node'].append(far if far is not None else 0.0)
                roc[current_method]['event_dr'].append(dr_env)
                roc[current_method]['event_dr_raw'].append(dr_raw if dr_raw is not None else dr_env)
                roc[current_method]['fp_clusters_total'].append(fp_clusters if fp_clusters is not None else 0)
        elif (not row or row[0] is None):
            # Blank row terminates the current detector block
            current_method = None
        i += 1

    # Sort each detector's points by FAR ascending to get a monotonic curve
    for method_key, data in roc.items():
        n = len(data['thresholds'])
        if n > 1:
            order = sorted(range(n), key=lambda j: data['fp_per_hour_per_node'][j])
            for k in data:
                data[k] = [data[k][j] for j in order]

    return roc


def collect_runs(workbook_path: Path) -> List[Dict]:
    """Load the workbook and parse out per-run metrics + ROC for every run."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    print(f"  Reading {workbook_path}")
    wb = load_workbook(workbook_path, data_only=True)

    # Configurations sheet -> list of runs (with metadata)
    runs = read_configurations(wb)
    if not runs:
        raise ValueError("No runs found in Configurations sheet")

    print(f"  Found {len(runs)} runs in Configurations sheet:")
    for r in runs:
        print(f"    {r['num_nodes']}n {int(r['snr_db'])}dB  ({r.get('origin')})")

    # Attach per-run metrics + ROC
    for r in runs:
        r['methods'] = read_per_run_metrics(wb, r)
        r['roc_sweep'] = read_per_run_roc(wb, r)
        n_roc_methods = len([m for m, d in r['roc_sweep'].items() if d.get('thresholds')])
        n_metric_methods = len([m for m, d in r['methods'].items() if d])
        print(f"    {r['num_nodes']}n {int(r['snr_db'])}dB: "
              f"{n_metric_methods}/5 detectors with metrics, "
              f"{n_roc_methods}/5 with ROC data")

    return runs


def run_label(run: Dict) -> str:
    """Compact label, e.g. '50n / 12dB' """
    nn = run.get('num_nodes', '?')
    snr = run.get('snr_db', '?')
    if isinstance(snr, float):
        snr = int(snr)
    return f"{nn}n / {snr}dB"


def get_metric(run: Dict, method: str, metric: str) -> Tuple[Optional[float], Optional[float]]:
    """Return (mean, std) for a metric, or (None, None) if missing."""
    return run.get('methods', {}).get(method, {}).get(metric, (None, None))


# ============================================================================
# Figure builders (identical to v1; data source is now the workbook)
# ============================================================================

def fig_01_headline_matrix(runs: List[Dict]) -> go.Figure:
    runs_sorted = sorted(runs, key=lambda r: (r.get('num_nodes', 0), -r.get('snr_db', 0)))
    config_labels = [run_label(r) for r in runs_sorted]
    detector_labels = [METHOD_LABELS[m] for m in ALL_METHODS]

    z, text = [], []
    for method_key in ALL_METHODS:
        z_row, text_row = [], []
        for run in runs_sorted:
            mean, std = get_metric(run, method_key, 'detection_rate')
            z_row.append(mean if mean is not None else float('nan'))
            if mean is None:
                text_row.append('—')
            elif std is not None and std > 0.05:
                text_row.append(f"{mean:.1f}%<br>±{std:.1f}")
            else:
                text_row.append(f"{mean:.1f}%")
        z.append(z_row)
        text.append(text_row)

    fig = go.Figure(data=go.Heatmap(
        z=z, x=config_labels, y=detector_labels,
        text=text, texttemplate='%{text}',
        textfont=dict(size=12, family='Arial', color='black'),
        colorscale=[
            [0.00, '#67000d'], [0.30, '#a50f15'], [0.50, '#fdae6b'],
            [0.70, '#fee08b'], [0.90, '#abdda4'], [1.00, '#1a9850'],
        ],
        zmin=0, zmax=100,
        colorbar=dict(title='Detection<br>Rate (%)', tickvals=[0, 25, 50, 75, 100]),
        hovertemplate='Detector: %{y}<br>Config: %{x}<br>Detection rate: %{z:.2f}%<extra></extra>',
    ))

    fig.update_layout(
        title='Detection Rate Matrix — All Detectors × All Configurations',
        xaxis_title='Configuration (network size / SNR)',
        yaxis_title='Detector',
        yaxis=dict(autorange='reversed'),
        height=420,
        **{k: v for k, v in FIG_LAYOUT.items() if k != 'margin'},
        margin=dict(l=110, r=120, t=80, b=80),
    )
    return fig


def fig_02_roc_envelope(runs: List[Dict]) -> go.Figure:
    runs_sorted = sorted(runs, key=lambda r: (-r.get('num_nodes', 0), -r.get('snr_db', 0)))
    n_runs = len(runs_sorted)
    cols = 2
    rows = (n_runs + cols - 1) // cols

    subplot_titles = [run_label(r) for r in runs_sorted]
    fig = make_subplots(
        rows=rows, cols=cols,
        subplot_titles=subplot_titles,
        shared_yaxes=False,
        horizontal_spacing=0.08,
        vertical_spacing=0.14,
    )

    FP_FLOOR = 1e-3
    showed_legend = set()

    for idx, run in enumerate(runs_sorted):
        row = idx // cols + 1
        col = idx % cols + 1
        roc = run.get('roc_sweep', {})

        for method_key in ALL_METHODS:
            data = roc.get(method_key)
            if not data:
                continue
            far = data.get('fp_per_hour_per_node') or []
            dr = data.get('event_dr') or []
            thresholds = data.get('thresholds') or []
            if not far or not dr:
                continue

            far_plot = [max(f, FP_FLOOR) for f in far]
            label = METHOD_LABELS[method_key]
            color = COLORS[method_key]

            show_in_legend = method_key not in showed_legend
            if show_in_legend:
                showed_legend.add(method_key)

            hovertext = [
                f"{label}<br>Threshold: {thr:.4g}<br>FAR: {f:.4g}/hr/node<br>DR: {d:.2f}%"
                for thr, f, d in zip(thresholds, far_plot, dr)
            ]

            fig.add_trace(
                go.Scatter(
                    x=far_plot, y=dr,
                    mode='lines+markers',
                    name=label,
                    legendgroup=method_key,
                    showlegend=show_in_legend,
                    line=dict(color=color, width=3 if method_key == 'proposed' else 2),
                    marker=dict(symbol=SYMBOLS[method_key],
                                size=8 if method_key == 'proposed' else 6,
                                color=color, line=dict(color='white', width=1)),
                    hovertext=hovertext, hoverinfo='text',
                ),
                row=row, col=col,
            )

            # Canonical operating point (threshold ≈ 1.0) — drawn as a hollow
            # circle ringing the underlying data point, so the reader sees
            # "this is the as-designed operating point on the curve".
            if thresholds:
                idx_canonical = min(range(len(thresholds)),
                                    key=lambda i: abs(thresholds[i] - 1.0))
                fig.add_trace(
                    go.Scatter(
                        x=[far_plot[idx_canonical]],
                        y=[dr[idx_canonical]],
                        mode='markers',
                        name=f"{label} canonical",
                        legendgroup=method_key,
                        showlegend=False,
                        marker=dict(symbol='circle-open', size=16,
                                    color=color,
                                    line=dict(color=color, width=2.5)),
                        hovertext=f"{label} canonical (thr ≈ 1.0)<br>"
                                 f"FAR: {far_plot[idx_canonical]:.4g}/hr/node<br>"
                                 f"DR: {dr[idx_canonical]:.2f}%",
                        hoverinfo='text',
                    ),
                    row=row, col=col,
                )

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            fig.update_xaxes(
                type='log',
                title_text='False-alarm cluster rate (/hr/node)',
                range=[math.log10(FP_FLOOR / 2), math.log10(1000)],
                gridcolor='#dddddd',
                row=r, col=c,
            )
            fig.update_yaxes(
                title_text='Event detection rate (%)',
                range=[-2, 105],
                gridcolor='#dddddd',
                row=r, col=c,
            )

    fig.update_layout(
        title=('ROC Envelope — TSNFA + 4 Classical Comparators '
               '<br><sub>Hollow circles mark canonical operating points (threshold ≈ 1.0). '
               'Curves are upper-envelope (Pareto frontier).</sub>'),
        height=400 * rows,
        **{k: v for k, v in FIG_LAYOUT.items() if k != 'margin'},
        margin=dict(l=70, r=30, t=120, b=60),
    )
    return fig


def fig_03_network_load(runs: List[Dict]) -> go.Figure:
    runs_sorted = sorted(runs, key=lambda r: (r.get('num_nodes', 0), -r.get('snr_db', 0)))
    config_labels = [run_label(r) for r in runs_sorted]

    fig = make_subplots(
        rows=1, cols=2,
        subplot_titles=('Per-Node Network Load (B/hr)',
                       'Total Mesh Load (kB/hr) = per-node × (N−1)'),
        horizontal_spacing=0.12,
    )

    for method_key in ALL_METHODS:
        per_node_vals, total_mesh_vals = [], []
        per_node_text, total_mesh_text = [], []

        for run in runs_sorted:
            mean, _ = get_metric(run, method_key, 'network_load_bytes_per_hour')
            n_sensor = max(0, (run.get('num_nodes', 1) or 1) - 1)
            per_node_vals.append(mean if mean is not None else 0)
            total_mesh = (mean * n_sensor / 1000.0) if mean is not None else 0
            total_mesh_vals.append(total_mesh)
            per_node_text.append(f"{mean:,.0f} B/hr" if mean is not None else '—')
            total_mesh_text.append(f"{total_mesh:,.1f} kB/hr" if mean is not None else '—')

        label = METHOD_LABELS[method_key]
        color = COLORS[method_key]

        fig.add_trace(
            go.Bar(
                x=config_labels, y=per_node_vals,
                name=label, legendgroup=method_key,
                marker_color=color,
                text=per_node_text, textposition='auto',
                hovertemplate=f"{label}<br>%{{x}}<br>Per-node: %{{y:,.0f}} B/hr<extra></extra>",
            ),
            row=1, col=1,
        )
        fig.add_trace(
            go.Bar(
                x=config_labels, y=total_mesh_vals,
                name=label, legendgroup=method_key,
                showlegend=False,
                marker_color=color,
                text=total_mesh_text, textposition='auto',
                hovertemplate=f"{label}<br>%{{x}}<br>Total mesh: %{{y:,.1f}} kB/hr<extra></extra>",
            ),
            row=1, col=2,
        )

    fig.update_yaxes(type='log', title_text='Bytes/hr (log)', row=1, col=1)
    fig.update_yaxes(type='log', title_text='kB/hr (log)', row=1, col=2)
    fig.update_xaxes(title_text='Configuration', row=1, col=1)
    fig.update_xaxes(title_text='Configuration', row=1, col=2)

    fig.update_layout(
        title=('Network Bandwidth Cost — Per-Node and Total Mesh '
               '<br><sub>Note log y-axis. TSNFA delivers full detection at fraction of comparator bandwidth.</sub>'),
        barmode='group',
        height=520,
        **{k: v for k, v in FIG_LAYOUT.items() if k != 'margin'},
        margin=dict(l=70, r=30, t=110, b=80),
    )
    return fig


def fig_04_snr_robustness(runs: List[Dict]) -> go.Figure:
    sizes = sorted({r['num_nodes'] for r in runs if r.get('num_nodes') is not None})
    snrs = sorted({r['snr_db'] for r in runs if r.get('snr_db') is not None})

    cols = max(1, len(sizes))
    fig = make_subplots(
        rows=1, cols=cols,
        subplot_titles=[f"{n}-node Network" for n in sizes],
        shared_yaxes=True,
        horizontal_spacing=0.06,
    )

    for col_idx, n_nodes in enumerate(sizes, start=1):
        showed_legend_here = (col_idx == 1)
        for method_key in ALL_METHODS:
            xs, ys, errors = [], [], []
            for snr in snrs:
                matching = [r for r in runs if r['num_nodes'] == n_nodes and r['snr_db'] == snr]
                if not matching:
                    continue
                mean, std = get_metric(matching[0], method_key, 'detection_rate')
                if mean is None:
                    continue
                xs.append(snr)
                ys.append(mean)
                errors.append(std if std is not None else 0)

            if not xs:
                continue

            label = METHOD_LABELS[method_key]
            color = COLORS[method_key]

            fig.add_trace(
                go.Scatter(
                    x=xs, y=ys,
                    error_y=dict(type='data', array=errors, visible=True),
                    mode='lines+markers',
                    name=label,
                    legendgroup=method_key,
                    showlegend=showed_legend_here,
                    line=dict(color=color, width=3 if method_key == 'proposed' else 2),
                    marker=dict(symbol=SYMBOLS[method_key],
                                size=12 if method_key == 'proposed' else 8,
                                color=color, line=dict(color='white', width=1)),
                    hovertemplate=f"{label}<br>SNR: %{{x}} dB<br>DR: %{{y:.2f}}%<extra></extra>",
                ),
                row=1, col=col_idx,
            )

    fig.update_xaxes(title_text='SNR (dB)', dtick=6, gridcolor='#dddddd')
    fig.update_yaxes(title_text='Event detection rate (%)',
                     range=[-5, 105], gridcolor='#dddddd', row=1, col=1)

    fig.update_layout(
        title=('SNR Robustness — Detection Rate vs SNR '
               '<br><sub>TSNFA holds 100% across all configurations; OS-CFAR collapses at 12 dB.</sub>'),
        height=520,
        **{k: v for k, v in FIG_LAYOUT.items() if k != 'margin'},
        margin=dict(l=70, r=30, t=110, b=70),
    )
    return fig


def fig_05_scaling(runs: List[Dict]) -> go.Figure:
    metrics_to_plot = [
        ('detection_rate',              'Detection Rate (%)',           False),
        ('event_precision',             'Event Precision (%)',          False),
        ('fp_clusters_outside',         'FP Clusters (count)',          True),
        ('false_alarm_rate_clusters',   'FAR clusters /hr/node',        False),
        ('latency_mean_ms',             'Mean Latency (ms)',            False),
        ('network_load_bytes_per_hour', 'Network Load (B/hr)',          True),
    ]
    snrs = sorted({r['snr_db'] for r in runs}, reverse=True)
    sizes = sorted({r['num_nodes'] for r in runs})

    rows, cols = 2, 3
    fig = make_subplots(
        rows=rows, cols=cols,
        subplot_titles=[m[1] for m in metrics_to_plot],
        horizontal_spacing=0.10,
        vertical_spacing=0.18,
    )

    showed_legend = set()
    for panel_idx, (metric_key, display, log_y) in enumerate(metrics_to_plot):
        row = panel_idx // cols + 1
        col = panel_idx % cols + 1

        for method_key in ALL_METHODS:
            for snr_idx, snr in enumerate(snrs):
                xs, ys = [], []
                for n_nodes in sizes:
                    matching = [r for r in runs
                                if r['num_nodes'] == n_nodes and r['snr_db'] == snr]
                    if not matching:
                        continue
                    mean, _ = get_metric(matching[0], method_key, metric_key)
                    if mean is None:
                        continue
                    xs.append(n_nodes)
                    ys.append(mean if mean > 0 or not log_y else 0.5)

                if not xs:
                    continue

                label = METHOD_LABELS[method_key]
                color = COLORS[method_key]
                dash = 'solid' if snr == max(snrs) else 'dash'

                tag = (method_key, snr)
                show_in_legend = tag not in showed_legend and panel_idx == 0
                if show_in_legend:
                    showed_legend.add(tag)

                trace_label = f"{label} ({int(snr)} dB)"

                fig.add_trace(
                    go.Scatter(
                        x=xs, y=ys,
                        mode='lines+markers',
                        name=trace_label,
                        legendgroup=f"{method_key}_{snr}",
                        showlegend=show_in_legend,
                        line=dict(color=color, width=2.5 if method_key == 'proposed' else 1.8,
                                  dash=dash),
                        marker=dict(symbol=SYMBOLS[method_key],
                                    size=10 if method_key == 'proposed' else 7,
                                    color=color, line=dict(color='white', width=1)),
                        hovertemplate=(
                            f"{trace_label}<br>"
                            f"N=%{{x}}<br>{display}=%{{y:.4g}}<extra></extra>"
                        ),
                    ),
                    row=row, col=col,
                )

        fig.update_xaxes(title_text='Network size (nodes)',
                         tickvals=sizes, gridcolor='#dddddd',
                         row=row, col=col)
        if log_y:
            fig.update_yaxes(type='log', title_text=display, gridcolor='#dddddd', row=row, col=col)
        else:
            fig.update_yaxes(title_text=display, gridcolor='#dddddd', row=row, col=col)

    fig.update_layout(
        title=('Network Scaling — Metric Behaviour as N Grows '
               '<br><sub>Solid lines: high SNR. Dashed lines: low SNR.</sub>'),
        height=720,
        **{k: v for k, v in FIG_LAYOUT.items() if k != 'margin'},
        margin=dict(l=70, r=30, t=110, b=70),
    )
    return fig


def fig_06_bandwidth_precision(runs: List[Dict]) -> go.Figure:
    fig = go.Figure()

    for method_key in ALL_METHODS:
        xs, ys, sizes_marker, hover_text = [], [], [], []
        for run in runs:
            load_mean, _ = get_metric(run, method_key, 'network_load_bytes_per_hour')
            prec_mean, _ = get_metric(run, method_key, 'event_precision')
            if load_mean is None or prec_mean is None:
                continue
            n_nodes = run.get('num_nodes', 10)
            marker_size = 12 + (n_nodes / 50) * 18
            xs.append(max(load_mean, 1))
            ys.append(prec_mean)
            sizes_marker.append(marker_size)
            hover_text.append(
                f"<b>{METHOD_LABELS[method_key]}</b><br>"
                f"Config: {run_label(run)}<br>"
                f"Network load: {load_mean:,.0f} B/hr<br>"
                f"Event precision: {prec_mean:.2f}%"
            )

        if not xs:
            continue

        label = METHOD_LABELS[method_key]
        color = COLORS[method_key]

        fig.add_trace(
            go.Scatter(
                x=xs, y=ys,
                mode='markers',
                name=label,
                marker=dict(symbol=SYMBOLS[method_key],
                            size=sizes_marker, color=color,
                            line=dict(color='white', width=2),
                            opacity=0.85),
                text=hover_text, hoverinfo='text',
            )
        )

    fig.add_annotation(
        x=math.log10(100), y=98,
        text='← Ideal: low bandwidth, high precision',
        showarrow=False,
        font=dict(size=11, color='#1f3864', style='italic'),
        bgcolor='rgba(255,255,255,0.7)',
        bordercolor='#1f3864', borderwidth=1, borderpad=4,
    )

    fig.update_xaxes(type='log', title_text='Network load per node (B/hr, log scale)',
                     gridcolor='#dddddd')
    fig.update_yaxes(title_text='Event precision (%)', range=[-5, 105], gridcolor='#dddddd')

    fig.update_layout(
        title=('Bandwidth-Precision Trade-off '
               '<br><sub>Marker size encodes network size. Top-left = ideal '
               '(precise + cheap).</sub>'),
        height=560,
        **{k: v for k, v in FIG_LAYOUT.items() if k != 'margin'},
        margin=dict(l=70, r=30, t=110, b=70),
    )
    return fig


def fig_07_3d_quality(runs: List[Dict]) -> go.Figure:
    """3D bubble plot: FAR x Network Load x Detection Rate per detector.

    Each detector contributes one bubble per (network size, SNR) configuration.
    Bubble size encodes Event Precision -- the "is this detection actually useful"
    measure that the three axes don't show. A "good" detector has a BIG bubble in
    the front-left-top corner: low FAR, low load, high DR, AND high precision.

    Visual encoding:
      - X axis  : FAR clusters/hr/node (log, low at left = good)
      - Y axis  : Network load B/hr (log, low at left = good)
      - Z axis  : Detection rate % (linear, high at top = good)
      - Bubble size  : Event Precision % (large = useful triggers; small = noise)
      - Bubble color : detector
      - Bubble shape : circle (consistent), but the screen-space disk is what
                       matters since the third dimension is the size encoding.

    Reference plane at DR = 100% gives the eye a "ceiling" so the height of each
    bubble below it reads as detection-rate deficit at a glance.
    """
    sizes_seen = sorted({r['num_nodes'] for r in runs if r.get('num_nodes') is not None})
    snrs_seen = sorted({r['snr_db'] for r in runs if r.get('snr_db') is not None})

    fig = go.Figure()

    FAR_FLOOR = 1e-3   # log axis floor for x
    LOAD_FLOOR = 1.0   # log axis floor for y

    # Bubble size mapping. Plotly Scatter3d marker.size is in pixels (default 6),
    # and we want a visually clear range from ~6 (low precision) to ~36 (high
    # precision). Apply a power transform so 50% -> noticeably bigger than 5%.
    PRECISION_MIN_PX = 6
    PRECISION_MAX_PX = 36
    def precision_to_size(prec_pct):
        if prec_pct is None:
            return PRECISION_MIN_PX
        # Clamp to [0, 100], take sqrt so low values are still visible
        p = max(0.0, min(100.0, prec_pct)) / 100.0
        return PRECISION_MIN_PX + (PRECISION_MAX_PX - PRECISION_MIN_PX) * (p ** 0.5)

    for method_key in ALL_METHODS:
        # Collect per-config points
        pts = []
        for n_nodes in sizes_seen:
            for snr in sorted(snrs_seen, reverse=True):
                matching = [r for r in runs
                            if r['num_nodes'] == n_nodes and r['snr_db'] == snr]
                if not matching:
                    continue
                run = matching[0]
                far_mean, _ = get_metric(run, method_key, 'false_alarm_rate_clusters')
                load_mean, _ = get_metric(run, method_key, 'network_load_bytes_per_hour')
                dr_mean, _ = get_metric(run, method_key, 'detection_rate')
                prec_mean, _ = get_metric(run, method_key, 'event_precision')
                if far_mean is None or load_mean is None or dr_mean is None:
                    continue
                pts.append({
                    'far':  max(far_mean, FAR_FLOOR),
                    'load': max(load_mean, LOAD_FLOOR),
                    'dr':   dr_mean,
                    'prec': prec_mean if prec_mean is not None else 0.0,
                    'n':    n_nodes,
                    'snr':  snr,
                })

        if not pts:
            continue

        label = METHOD_LABELS[method_key]
        color = COLORS[method_key]
        is_proposed = (method_key == 'proposed')

        xs = [p['far'] for p in pts]
        ys = [p['load'] for p in pts]
        zs = [p['dr'] for p in pts]
        sizes_px = [precision_to_size(p['prec']) for p in pts]

        hover_texts = [
            f"<b>{label}</b><br>"
            f"Config: {p['n']}n / {int(p['snr'])} dB<br>"
            f"FAR: {p['far']:.3g}/hr/node<br>"
            f"Load: {p['load']:,.0f} B/hr<br>"
            f"DR: {p['dr']:.2f}%<br>"
            f"<b>Precision: {p['prec']:.2f}%</b> (bubble size)"
            for p in pts
        ]

        fig.add_trace(go.Scatter3d(
            x=xs, y=ys, z=zs,
            mode='markers',
            name=label,
            legendgroup=method_key,
            marker=dict(
                size=sizes_px,
                color=color,
                opacity=0.75 if not is_proposed else 0.85,
                line=dict(color='rgba(20,20,20,0.4)', width=1),
                # sizemode 'diameter' makes the size value the actual diameter in pixels
                sizemode='diameter',
            ),
            text=hover_texts,
            hoverinfo='text',
        ))

    # Reference plane at DR = 100% -- a faint translucent surface that gives the
    # eye a visual anchor. Bubbles below the plane visibly show their DR deficit.
    far_plane = [FAR_FLOOR, 200]
    load_plane = [LOAD_FLOOR, 2e6]
    fig.add_trace(go.Surface(
        x=far_plane,
        y=load_plane,
        z=[[100, 100], [100, 100]],
        colorscale=[[0, 'rgba(46, 204, 113, 0.08)'], [1, 'rgba(46, 204, 113, 0.08)']],
        showscale=False,
        name='DR = 100% reference',
        hoverinfo='skip',
        showlegend=False,
    ))

    # Bubble-size legend (a separate trace at the corner, not on a real data
    # point, that shows the precision -> bubble-size mapping)
    legend_precisions = [10, 50, 100]
    legend_xs = [FAR_FLOOR * 1.3] * len(legend_precisions)
    legend_ys = [LOAD_FLOOR * 1.3] * len(legend_precisions)
    legend_zs = [10, 50, 100][:len(legend_precisions)]  # bottom-left corner stack
    legend_sizes = [precision_to_size(p) for p in legend_precisions]
    fig.add_trace(go.Scatter3d(
        x=legend_xs, y=legend_ys, z=legend_zs,
        mode='markers+text',
        marker=dict(
            size=legend_sizes,
            color='rgba(150,150,150,0.35)',
            line=dict(color='rgba(80,80,80,0.6)', width=1),
            sizemode='diameter',
        ),
        text=[f'  {p}% prec' for p in legend_precisions],
        textposition='middle right',
        textfont=dict(size=10, color='#444'),
        hoverinfo='skip',
        showlegend=False,
        name='Precision legend',
    ))

    # Layout
    fig.update_layout(
        title=('Detector Quality in 3D — FAR × Network Load × Detection Rate '
               '<br><sub>Bubble size = Event Precision. A "good" detector is a '
               '<b>big bubble in the front-left-top corner</b> '
               '(low FAR, low load, high DR, high precision). Drag to rotate.</sub>'),
        scene=dict(
            xaxis=dict(
                title='FAR clusters /hr/node (log)',
                type='log',
                gridcolor='#dddddd',
                backgroundcolor='#fafbfc',
                showbackground=True,
                range=[math.log10(FAR_FLOOR / 2), math.log10(200)],
                # Reverse axis so "low FAR / good" ends up on the LEFT (near the viewer)
                autorange='reversed',
            ),
            yaxis=dict(
                title='Network load B/hr (log)',
                type='log',
                gridcolor='#dddddd',
                backgroundcolor='#fafbfc',
                showbackground=True,
                range=[math.log10(LOAD_FLOOR / 2), math.log10(2e6)],
                # Reverse so "low load / good" is at the LEFT
                autorange='reversed',
            ),
            zaxis=dict(
                title='Detection rate (%)',
                gridcolor='#dddddd',
                backgroundcolor='#fafbfc',
                showbackground=True,
                range=[0, 105],
            ),
            # Camera looks at the cube from front-upper-right, slightly elevated.
            # With both X and Y reversed, the "good corner" (low FAR, low load,
            # high DR) is now front-left-top from the camera's perspective.
            camera=dict(
                eye=dict(x=1.7, y=1.7, z=1.0),
                up=dict(x=0, y=0, z=1),
                center=dict(x=0, y=0, z=-0.1),
            ),
            aspectmode='cube',
        ),
        height=750,
        **{k: v for k, v in FIG_LAYOUT.items() if k != 'margin'},
        margin=dict(l=10, r=10, t=110, b=10),
    )

    return fig



# ============================================================================
# Pre-computed sheet readers (for the rich tables on the data page)
# ============================================================================

def read_headline_metrics(wb) -> Dict:
    """Parse Headline_Metrics sheet -> {detector: {metric: {config_label: value}}}.

    Returns a nested dict so the renderer can emit a table grouped by detector.
    Config labels are like '50n / 12dB', matching the column headers in the sheet.
    """
    if 'Headline_Metrics' not in wb.sheetnames:
        return {}

    ws = wb['Headline_Metrics']
    # Row 3 is the header. Columns 3+ are config labels.
    headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    config_cols = []
    for col_idx, h in enumerate(headers, start=1):
        if isinstance(h, str) and 'n /' in h and 'dB' in h:
            config_cols.append((h, col_idx))

    result = {}
    for row in range(4, ws.max_row + 1):
        detector = ws.cell(row=row, column=1).value
        metric = ws.cell(row=row, column=2).value
        if not detector or not metric:
            continue
        if detector not in result:
            result[detector] = {}
        if metric not in result[detector]:
            result[detector][metric] = {}
        for cfg_label, col in config_cols:
            v = ws.cell(row=row, column=col).value
            result[detector][metric][cfg_label] = v
    return result


def read_network_load_comparison(wb) -> Dict:
    """Parse the per-node section of Network_Load_Comparison.

    Returns {detector: {config_label: bytes_per_hour}}.
    """
    if 'Network_Load_Comparison' not in wb.sheetnames:
        return {}

    ws = wb['Network_Load_Comparison']
    # Row 5 is the header for the per-node sub-table.
    headers = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column + 1)]
    config_cols = []
    for col_idx, h in enumerate(headers, start=1):
        if isinstance(h, str) and 'n /' in h and 'dB' in h:
            config_cols.append((h, col_idx))

    result = {}
    # Per-node section runs rows 6-10
    for row in range(6, 11):
        detector = ws.cell(row=row, column=1).value
        if not detector:
            continue
        result[detector] = {}
        for cfg_label, col in config_cols:
            result[detector][cfg_label] = ws.cell(row=row, column=col).value
    return result


def read_snr_robustness(wb) -> Dict:
    """Parse the SNR_Robustness sheet.

    Returns {network_size: {detector: {'18dB': v, '12dB': v, 'drop': v}}}.
    """
    if 'SNR_Robustness' not in wb.sheetnames:
        return {}

    ws = wb['SNR_Robustness']
    result = {}

    # Find size sections by scanning column 1 for "Network Size: N nodes" rows
    rows = list(ws.iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        v = rows[i][0] if rows[i] else None
        if isinstance(v, str) and v.startswith('Network Size:'):
            # extract N
            m = re.search(r'(\d+)\s*nodes', v)
            if m:
                size_key = int(m.group(1))
                result[size_key] = {}
                # next row is header (Detector | 18 dB | 12 dB | SNR Drop)
                # then 5 detector rows
                for det_row_idx in range(i + 2, i + 7):
                    if det_row_idx >= len(rows):
                        break
                    r = rows[det_row_idx]
                    if not r or not r[0]:
                        break
                    detector = r[0]
                    result[size_key][detector] = {
                        '18dB': r[1],
                        '12dB': r[2],
                        'drop': r[3],
                    }
        i += 1
    return result


def read_algorithm_parameters(wb) -> List[Dict]:
    """Parse Configurations sheet for algorithm-parameter columns only.

    Returns a list of {Run, TSNFA params, Lipski params, etc.} dicts.
    """
    if 'Configurations' not in wb.sheetnames:
        return []

    ws = wb['Configurations']
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    rows_out = []
    for row in range(2, ws.max_row + 1):
        d = {}
        for col_idx, h in enumerate(headers, start=1):
            if h is None:
                continue
            d[h] = ws.cell(row=row, column=col_idx).value
        if d.get('Run'):
            rows_out.append(d)
    return rows_out


# ============================================================================
# HTML table renderers
# ============================================================================

def _fmt_num(v, decimals=2, fallback='—'):
    """Format a number with given decimals, or fallback if None."""
    if v is None:
        return fallback
    try:
        return f"{float(v):,.{decimals}f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_pct(v, fallback='—'):
    """Format a percentage value."""
    if v is None:
        return fallback
    try:
        return f"{float(v):.2f}%"
    except (TypeError, ValueError):
        return str(v)


def _fmt_int(v, fallback='—'):
    if v is None:
        return fallback
    try:
        return f"{int(v):,}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_bytes_per_hr(v, fallback='—'):
    if v is None:
        return fallback
    try:
        x = float(v)
        if x < 10:
            return f"{x:.2f} B/hr"
        if x < 1000:
            return f"{x:.0f} B/hr"
        if x < 1e6:
            return f"{x/1000:.1f} kB/hr"
        return f"{x/1e6:.2f} MB/hr"
    except (TypeError, ValueError):
        return str(v)


def render_run_inventory_table(runs: List[Dict]) -> str:
    rows = []
    for r in sorted(runs, key=lambda x: (x.get('num_nodes', 0), -x.get('snr_db', 0))):
        rows.append(
            f"<tr>"
            f"<td><strong>{run_label(r)}</strong></td>"
            f"<td><code>{r.get('origin', '?')}</code></td>"
            f"<td>{r.get('num_nodes')}</td>"
            f"<td>{int(r.get('snr_db'))}</td>"
            f"<td>{_fmt_num(r.get('duration_hours'), 1)}h</td>"
            f"<td>{r.get('num_mc_runs')}</td>"
            f"<td>{_fmt_int(r.get('total_events'))}</td>"
            f"<td>{r.get('preset')}</td>"
            f"</tr>"
        )
    return f"""<table class="data-table">
<thead><tr>
  <th>Configuration</th>
  <th>Origin</th>
  <th>N nodes</th>
  <th>SNR (dB)</th>
  <th>Duration</th>
  <th>MC Runs</th>
  <th>Total Events</th>
  <th>Preset</th>
</tr></thead>
<tbody>
  {chr(10).join(rows)}
</tbody></table>"""


def render_algorithm_params_table(runs: List[Dict]) -> str:
    if not runs:
        return ""
    # Algorithm parameters are the same across runs (preset-driven), so just
    # show the values from the first run as the canonical settings.
    r = runs[0]
    sections = [
        ("TSNFA", [
            ("γ_d (Defence-1 threshold)", r.get('tsnfa_gamma_d')),
            ("γ_a (Defence-2 threshold)", r.get('tsnfa_gamma_a')),
            ("ζ (CFAR multiplier)",       r.get('tsnfa_zeta')),
        ]),
        ("Lipski FFT", [
            ("k (sigma multiplier)",      r.get('lipski_k')),
            ("N_bins_min (min consecutive bins)", r.get('Lipski n_bins_min') if 'Lipski n_bins_min' in r else None),
        ]),
        ("CA-CFAR", [
            ("N_ref (reference cells)",   r.get('CA-CFAR n_ref') if 'CA-CFAR n_ref' in r else None),
            ("P_fa (false-alarm prob)",   r.get('cacfar_p_fa')),
        ]),
        ("OS-CFAR", [
            ("N_ref (reference cells)",   r.get('OS-CFAR n_ref') if 'OS-CFAR n_ref' in r else None),
            ("k_rank (rank statistic)",   r.get('oscfar_k_rank')),
            ("P_fa (false-alarm prob)",   r.get('oscfar_p_fa')),
        ]),
        ("CUSUM", [
            ("α_fa (false-alarm rate)",   r.get('cusum_alpha_fa')),
        ]),
    ]

    rows = []
    for det_name, params in sections:
        for i, (param, value) in enumerate(params):
            det_cell = f'<td rowspan="{len(params)}"><strong>{det_name}</strong></td>' if i == 0 else ''
            value_str = '—' if value is None else (
                f"{value:.4g}" if isinstance(value, (int, float)) else str(value)
            )
            rows.append(f"<tr>{det_cell}<td>{param}</td><td><code>{value_str}</code></td></tr>")

    return f"""<table class="data-table">
<thead><tr><th>Detector</th><th>Parameter</th><th>Canonical Value</th></tr></thead>
<tbody>{chr(10).join(rows)}</tbody>
</table>"""


def render_headline_table(headline: Dict, runs: List[Dict]) -> str:
    """A 5-detector x N-config table for a chosen set of headline metrics."""
    if not headline:
        return "<p><em>(Headline metrics not available)</em></p>"

    # Pick the metrics most worth showing on the front page
    metrics = [
        ('Detection Rate (%)',   '%.2f%%',   None),
        ('Event Precision (%)',  '%.2f%%',   None),
        ('FP Clusters',          '%d',       None),
        ('Network Load (B/hr)',  None,       'bytes_per_hr'),
    ]

    # Get configs in display order
    configs = sorted({lab for det, m in headline.items() for met, vals in m.items()
                      for lab in vals.keys()},
                     key=lambda s: (int(re.search(r'(\d+)n', s).group(1)),
                                    -int(re.search(r'(\d+)dB', s).group(1))))

    detectors = ['TSNFA', 'Lipski', 'CA-CFAR', 'OS-CFAR', 'CUSUM']
    rows = []

    for metric_name, fmt, special in metrics:
        # Each metric gets its own header row
        rows.append(
            f'<tr class="metric-header"><td colspan="{len(configs) + 1}">'
            f'<strong>{metric_name}</strong></td></tr>'
        )
        for det in detectors:
            cells = [f"<td>{det}</td>"]
            for cfg in configs:
                v = headline.get(det, {}).get(metric_name, {}).get(cfg)
                if v is None:
                    cells.append("<td>—</td>")
                elif special == 'bytes_per_hr':
                    cells.append(f"<td>{_fmt_bytes_per_hr(v)}</td>")
                elif fmt:
                    try:
                        cells.append(f"<td>{fmt % float(v)}</td>")
                    except (TypeError, ValueError):
                        cells.append(f"<td>{v}</td>")
                else:
                    cells.append(f"<td>{v}</td>")
            cls = ' class="proposed-row"' if det == 'TSNFA' else ''
            rows.append(f"<tr{cls}>{''.join(cells)}</tr>")

    header_cells = ''.join(f"<th>{c}</th>" for c in configs)
    return f"""<table class="data-table">
<thead><tr><th>Detector</th>{header_cells}</tr></thead>
<tbody>{chr(10).join(rows)}</tbody></table>"""


def render_network_load_table(network: Dict) -> str:
    """Per-node network-load table with computed total mesh and ratio columns."""
    if not network:
        return "<p><em>(Network load data not available)</em></p>"

    detectors = ['TSNFA', 'Lipski', 'CA-CFAR', 'OS-CFAR', 'CUSUM']
    configs = sorted({c for d in network.values() for c in d.keys()},
                     key=lambda s: (int(re.search(r'(\d+)n', s).group(1)),
                                    -int(re.search(r'(\d+)dB', s).group(1))))

    rows = []
    # Compute TSNFA per-config values for ratios
    tsnfa_vals = {cfg: network.get('TSNFA', {}).get(cfg) for cfg in configs}

    for det in detectors:
        cls = ' class="proposed-row"' if det == 'TSNFA' else ''
        cells = [f"<td>{det}</td>"]
        for cfg in configs:
            v = network.get(det, {}).get(cfg)
            n_nodes = int(re.search(r'(\d+)n', cfg).group(1))
            n_sensor = max(0, n_nodes - 1)
            if v is None:
                cells.append('<td>—</td><td>—</td><td>—</td>')
                continue
            per_node = float(v)
            total_mesh_kb = per_node * n_sensor / 1000.0
            tref = tsnfa_vals.get(cfg)
            if tref and tref > 0:
                ratio = per_node / float(tref)
                ratio_str = f"{ratio:,.1f}×" if ratio >= 1 else f"{ratio:.3f}×"
            else:
                ratio_str = '—'
            cells.append(f'<td>{_fmt_bytes_per_hr(per_node)}</td>'
                         f'<td>{total_mesh_kb:,.1f} kB/hr</td>'
                         f'<td>{ratio_str}</td>')
        rows.append(f"<tr{cls}>{''.join(cells)}</tr>")

    # Header: 3 sub-columns per config (per-node, total mesh, ratio vs TSNFA)
    config_th = ''.join(
        f'<th colspan="3" class="config-group">{c}</th>' for c in configs
    )
    sub_th = ''.join('<th>Per-node</th><th>Total mesh</th><th>vs TSNFA</th>'
                     for _ in configs)

    return f"""<table class="data-table data-table-compact">
<thead>
  <tr><th rowspan="2">Detector</th>{config_th}</tr>
  <tr>{sub_th}</tr>
</thead>
<tbody>{chr(10).join(rows)}</tbody></table>"""


def render_snr_robustness_table(snr_data: Dict) -> str:
    if not snr_data:
        return "<p><em>(SNR robustness data not available)</em></p>"

    detectors = ['TSNFA', 'Lipski', 'CA-CFAR', 'OS-CFAR', 'CUSUM']
    sizes = sorted(snr_data.keys())

    rows = []
    for det in detectors:
        cls = ' class="proposed-row"' if det == 'TSNFA' else ''
        cells = [f"<td>{det}</td>"]
        for size in sizes:
            d = snr_data[size].get(det, {})
            dr_18 = d.get('18dB')
            dr_12 = d.get('12dB')
            drop = d.get('drop')
            drop_str = '—'
            drop_cls = ''
            if drop is not None:
                try:
                    dval = float(drop)
                    drop_str = f"{dval:+.2f}" if abs(dval) > 0.01 else "≈0"
                    if dval > 30:
                        drop_cls = ' class="cell-bad"'
                    elif dval > 10:
                        drop_cls = ' class="cell-warn"'
                    elif abs(dval) < 1:
                        drop_cls = ' class="cell-good"'
                except (TypeError, ValueError):
                    pass
            cells.append(
                f'<td>{_fmt_pct(dr_18)}</td>'
                f'<td>{_fmt_pct(dr_12)}</td>'
                f'<td{drop_cls}>{drop_str}</td>'
            )
        rows.append(f"<tr{cls}>{''.join(cells)}</tr>")

    size_th = ''.join(
        f'<th colspan="3" class="config-group">N = {s}</th>' for s in sizes
    )
    sub_th = ''.join('<th>18 dB</th><th>12 dB</th><th>Δ%</th>' for _ in sizes)

    return f"""<table class="data-table data-table-compact">
<thead>
  <tr><th rowspan="2">Detector</th>{size_th}</tr>
  <tr>{sub_th}</tr>
</thead>
<tbody>{chr(10).join(rows)}</tbody></table>"""


# ============================================================================
# Index page (rich report layout with tables + inline figures)
# ============================================================================

INDEX_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>TSNFA Monte Carlo — Data Page</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
    line-height: 1.65;
    color: #2c2c2c;
    background: #fafbfc;
    margin: 0;
    padding: 0;
  }}
  .container {{
    max-width: 1080px;
    margin: 0 auto;
    padding: 40px 30px 80px;
  }}
  header.page-header {{
    border-bottom: 3px solid #1f3864;
    margin-bottom: 32px;
    padding-bottom: 20px;
  }}
  h1 {{
    color: #1f3864;
    font-size: 2.0em;
    font-weight: 700;
    margin: 0 0 8px 0;
  }}
  h2 {{
    color: #1f3864;
    margin-top: 56px;
    border-bottom: 1px solid #d0d7de;
    padding-bottom: 8px;
    font-size: 1.45em;
  }}
  h3 {{
    color: #1f3864;
    margin-top: 28px;
    font-size: 1.15em;
  }}
  p.subtitle {{
    color: #666;
    font-size: 1.1em;
    margin: 0;
  }}
  .meta-bar {{
    display: flex;
    gap: 24px;
    flex-wrap: wrap;
    background: #f1f4f9;
    border-left: 4px solid #1f3864;
    padding: 14px 20px;
    margin: 24px 0;
    font-size: 0.95em;
  }}
  .meta-item strong {{ color: #1f3864; }}
  /* Data tables */
  table.data-table {{
    width: 100%;
    border-collapse: collapse;
    margin: 16px 0 24px;
    background: white;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    overflow: hidden;
    font-size: 0.92em;
  }}
  table.data-table th, table.data-table td {{
    padding: 8px 12px;
    text-align: left;
    border-bottom: 1px solid #ebeef2;
  }}
  table.data-table th {{
    background: #1f3864;
    color: white;
    font-weight: 600;
    font-size: 0.9em;
    text-align: center;
  }}
  table.data-table th.config-group {{
    border-bottom: 1px solid #2d4d8a;
    background: #2d4d8a;
  }}
  table.data-table tbody tr:hover {{ background: #f6f9fc; }}
  table.data-table tr:nth-child(even):not(.proposed-row):not(.metric-header) {{ background: #f9fafc; }}
  table.data-table td:first-child {{ font-weight: 500; }}
  /* TSNFA highlight row */
  table.data-table tr.proposed-row {{
    background: #e8f5e9 !important;
    font-weight: 500;
  }}
  table.data-table tr.proposed-row td {{
    border-bottom-color: #c8e6c9;
  }}
  /* Metric-header rows for grouped headline table */
  table.data-table tr.metric-header td {{
    background: #f1f4f9;
    color: #1f3864;
    padding: 6px 12px;
    font-size: 0.92em;
  }}
  /* Drop-cell coloring for SNR Robustness */
  table.data-table td.cell-bad {{
    background: #ffebee;
    color: #c62828;
    font-weight: 600;
  }}
  table.data-table td.cell-warn {{
    background: #fff8e1;
    color: #ef6c00;
    font-weight: 600;
  }}
  table.data-table td.cell-good {{
    background: #e8f5e9;
    color: #2e7d32;
    font-weight: 600;
  }}
  /* Compact table variant */
  table.data-table-compact th, table.data-table-compact td {{
    padding: 6px 10px;
    font-size: 0.88em;
  }}
  /* Detector cards */
  .detector-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 14px;
    margin: 18px 0;
  }}
  .detector-card {{
    background: white;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    padding: 14px 16px;
    border-left: 4px solid #1f3864;
  }}
  .detector-card.proposed {{ border-left-color: #2ecc71; }}
  .detector-card h4 {{
    color: #1f3864;
    margin: 0 0 6px 0;
    font-size: 1.02em;
  }}
  .detector-card.proposed h4 {{ color: #2ecc71; }}
  .detector-card p {{ margin: 0; font-size: 0.92em; color: #555; }}
  /* Figure embeds */
  .figure-section {{
    background: white;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    margin: 18px 0;
    overflow: hidden;
  }}
  .figure-section iframe {{
    width: 100%;
    border: 0;
    display: block;
  }}
  .figure-caption {{
    padding: 10px 16px;
    background: #f6f9fc;
    border-top: 1px solid #ebeef2;
    font-size: 0.88em;
    color: #555;
  }}
  .figure-caption strong {{ color: #1f3864; }}
  .figure-caption a {{ color: #1f3864; }}
  /* Notes */
  .notes {{
    background: #fffbe6;
    border-left: 4px solid #faad14;
    padding: 14px 20px;
    margin: 20px 0;
    border-radius: 4px;
    font-size: 0.95em;
  }}
  .notes ul {{ margin: 4px 0 0 0; padding-left: 20px; }}
  .notes li {{ margin: 4px 0; }}
  .key-finding {{
    background: #e8f5e9;
    border-left: 4px solid #2ecc71;
    padding: 14px 20px;
    margin: 16px 0;
    border-radius: 4px;
    font-size: 0.97em;
  }}
  .key-finding strong {{ color: #1b5e20; }}
  footer {{
    margin-top: 80px;
    padding-top: 20px;
    border-top: 1px solid #d0d7de;
    color: #888;
    font-size: 0.85em;
    text-align: center;
  }}
  code {{
    background: #f1f4f9;
    padding: 2px 6px;
    border-radius: 3px;
    font-size: 0.92em;
    font-family: 'Consolas', 'Monaco', monospace;
  }}
  /* Sticky table of contents */
  nav.toc {{
    background: #f1f4f9;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    padding: 12px 18px;
    margin: 24px 0;
    font-size: 0.92em;
  }}
  nav.toc ul {{
    margin: 4px 0 0 0;
    padding-left: 20px;
    columns: 2;
  }}
  nav.toc li {{ margin: 3px 0; }}
  nav.toc a {{ color: #1f3864; text-decoration: none; }}
  nav.toc a:hover {{ text-decoration: underline; }}
  @media (max-width: 600px) {{
    nav.toc ul {{ columns: 1; }}
    .container {{ padding: 20px 14px; }}
  }}
</style>
</head>
<body>
<div class="container">

<header class="page-header">
  <h1>TSNFA Monte Carlo Simulation — Data Page</h1>
  <p class="subtitle">Comprehensive comparison of TSNFA against four classical detection
  algorithms across {n_runs} Monte Carlo configurations.</p>
</header>

<div class="meta-bar">
  <div class="meta-item"><strong>Generated:</strong> {generated_at}</div>
  <div class="meta-item"><strong>Source:</strong> <code>{workbook_basename}</code></div>
  <div class="meta-item"><strong>Configurations:</strong> {config_summary}</div>
</div>

<nav class="toc">
  <strong>Contents</strong>
  <ul>
    <li><a href="#runs">1. Run Inventory</a></li>
    <li><a href="#detectors">2. Detector Pool</a></li>
    <li><a href="#parameters">3. Algorithm Parameters</a></li>
    <li><a href="#headline">4. Headline Results</a></li>
    <li><a href="#roc">5. ROC Analysis</a></li>
    <li><a href="#bandwidth">6. Network Bandwidth</a></li>
    <li><a href="#snr">7. SNR Robustness</a></li>
    <li><a href="#scaling">8. Network Scaling</a></li>
    <li><a href="#tradeoff">9. Bandwidth–Precision</a></li>
    <li><a href="#3d">10. 3D Quality Space</a></li>
    <li><a href="#methodology">11. Methodology</a></li>
  </ul>
</nav>

<h2 id="runs">1. Run Inventory</h2>
<p>Four Monte Carlo configurations were run, factorial across two network sizes
(10 and 50 nodes) and two SNR levels (12 dB and 18 dB). Each configuration was
simulated for 24 hours of synthetic time and replicated five times with
independent random seeds for variance estimation.</p>
{run_inventory_table}

<h2 id="detectors">2. Detector Pool</h2>
<p>Five detection algorithms were evaluated. TSNFA is the proposed cascade.
Lipski FFT, CA-CFAR, OS-CFAR, and CUSUM are classical baselines drawn from
radar and statistical signal processing literature, used here at their
canonical published settings to allow a fair comparison.</p>
<div class="detector-grid">
  <div class="detector-card proposed">
    <h4>TSNFA (proposed)</h4>
    <p>Two-stage cascade: spectral consistency (FFT-binned, Defence-1) followed
    by median-CFAR rank statistic (Defence-2). Designed for low-SNR seismic
    monitoring with strict bandwidth constraints.</p>
  </div>
  <div class="detector-card">
    <h4>Lipski FFT</h4>
    <p>Per-bin energy detector firing when bin power exceeds μ + kσ for a
    minimum number of consecutive bins. Canonical k = 3 (Lipski et al.).</p>
  </div>
  <div class="detector-card">
    <h4>CA-CFAR</h4>
    <p>Cell-averaging Constant False Alarm Rate detector using a sliding mean
    of N reference cells. Threshold multiplier α = 7.71 derived from
    P_fa = 10⁻³ via the Finn-Johnson formula.</p>
  </div>
  <div class="detector-card">
    <h4>OS-CFAR</h4>
    <p>Order-statistic CFAR using the 75th-percentile (k = 24 of N = 32) of the
    reference cells. Threshold multiplier α = 37.33 from
    P_fa = 10⁻³ (Rohling 1983).</p>
  </div>
  <div class="detector-card">
    <h4>CUSUM</h4>
    <p>Tartakovsky's bounded-memory cumulative-sum change detector with
    α_fa = 10⁻⁵ and a finite window K_end = 100 to prevent unbounded
    accumulation in stationary noise.</p>
  </div>
</div>

<h2 id="parameters">3. Algorithm Parameters</h2>
<p>The canonical settings used across all four runs. These are fixed
by-design parameters drawn from the published literature, not tuned to the
data. Identical settings ensure detectors are compared at their published
operating points rather than at empirically-optimized configurations.</p>
{algorithm_params_table}

<h2 id="headline">4. Headline Results</h2>
<p>The central result table. Each detector evaluated across all four
configurations on the four most diagnostic metrics: detection rate, event
precision, false-positive cluster count, and per-node network load.</p>
{headline_table}
<div class="key-finding">
  <strong>Key finding:</strong> TSNFA achieves 99.9–100% detection rate with 100%
  precision (zero false-positive clusters) across all four configurations. No
  classical comparator achieves both. Lipski and CA-CFAR achieve high DR
  (~99.7–100%) but with precision below 3%. OS-CFAR shows severe SNR-dependent
  collapse. CUSUM shows moderate brittleness.
</div>

<div class="figure-section">
  <iframe src="fig_01_headline_matrix.html" height="450"></iframe>
  <div class="figure-caption">
    <strong>Figure 1.</strong> Detection rate matrix — heat-map overview of
    detection rate per detector × per configuration. TSNFA's row is uniformly
    deep green; OS-CFAR's row shows the SNR-dependent collapse.
    <a href="fig_01_headline_matrix.html">Open standalone →</a>
  </div>
</div>

<h2 id="roc">5. ROC Analysis</h2>
<p>For each detector, the simulator swept a threshold multiplier in
post-processing (using the saved per-frame strengths) to trace out the ROC
upper envelope. Hollow circles mark the canonical operating point — the
threshold value that produces the headline metrics in Section 4.</p>
<div class="figure-section">
  <iframe src="fig_02_roc_envelope.html" height="850"></iframe>
  <div class="figure-caption">
    <strong>Figure 2.</strong> ROC envelopes per configuration. Toggle detectors
    via the legend; drag to zoom into the upper-left corner.
    <a href="fig_02_roc_envelope.html">Open standalone →</a>
  </div>
</div>

<h2 id="bandwidth">6. Network Bandwidth</h2>
<p>Network load per detector, expressed as bytes per hour each sensor node
transmits. Total mesh load is per-node × (N − 1), where N − 1 is the number of
sensor nodes (one node is the sink). The "vs TSNFA" column shows the bandwidth
ratio, revealing how much more traffic each comparator generates at the same
operating point.</p>
{network_load_table}
<div class="key-finding">
  <strong>Key finding:</strong> At 50n × 12 dB, Lipski generates 1025× more
  per-node bandwidth than TSNFA, CA-CFAR generates 617×, and CUSUM generates
  49×. OS-CFAR's lower bandwidth (146 B/hr vs TSNFA's 1160 B/hr) is a
  consequence of its 96.8% miss rate — a silent detector consumes little
  bandwidth.
</div>
<div class="figure-section">
  <iframe src="fig_03_network_load.html" height="560"></iframe>
  <div class="figure-caption">
    <strong>Figure 3.</strong> Per-node and total mesh bandwidth across
    configurations. Log y-axis spans three orders of magnitude.
    <a href="fig_03_network_load.html">Open standalone →</a>
  </div>
</div>

<h2 id="snr">7. SNR Robustness</h2>
<p>How much each detector's detection rate degrades when SNR drops from 18 dB
to 12 dB. Robust detectors show ≈ 0 percentage-point drop; brittle detectors
collapse. Cells highlighted: red for severe brittleness (>30 pp drop), amber
for moderate (10–30 pp), green for stability (&lt; 1 pp).</p>
{snr_robustness_table}
<div class="figure-section">
  <iframe src="fig_04_snr_robustness.html" height="560"></iframe>
  <div class="figure-caption">
    <strong>Figure 4.</strong> Detection rate vs SNR, paired panels per network
    size. TSNFA holds a flat 100% line; OS-CFAR shows the steep collapse.
    <a href="fig_04_snr_robustness.html">Open standalone →</a>
  </div>
</div>

<h2 id="scaling">8. Network Scaling</h2>
<p>Behaviour of each metric as network size grows from 10 to 50 nodes.
Detection-quality metrics (DR, precision) are per-node algorithmic properties
and should be invariant in N. Cost metrics (FP cluster count, network load)
scale linearly because each node operates independently.</p>
<div class="figure-section">
  <iframe src="fig_05_scaling.html" height="760"></iframe>
  <div class="figure-caption">
    <strong>Figure 5.</strong> Six-metric scaling grid. Solid lines: high SNR.
    Dashed lines: low SNR.
    <a href="fig_05_scaling.html">Open standalone →</a>
  </div>
</div>

<h2 id="tradeoff">9. Bandwidth–Precision Trade-off</h2>
<p>Each detector × configuration plotted as a single point in 2D (network load,
event precision) space. The upper-left corner is ideal: high precision and low
bandwidth. Marker size encodes network size.</p>
<div class="figure-section">
  <iframe src="fig_06_bandwidth_precision.html" height="600"></iframe>
  <div class="figure-caption">
    <strong>Figure 6.</strong> Bandwidth-precision trade-off. TSNFA occupies
    the upper-left "precise + cheap" corner alone.
    <a href="fig_06_bandwidth_precision.html">Open standalone →</a>
  </div>
</div>

<h2 id="3d">10. 3D Quality Space</h2>
<p>Each detector × configuration as a bubble in three dimensions
(FAR × network load × detection rate). Bubble size encodes event precision
(big = useful triggers, small = noise). A "good" detector is a big bubble in
the front-left-top corner: low FAR, low load, high DR, high precision.</p>
<div class="figure-section">
  <iframe src="fig_07_3d_quality.html" height="800"></iframe>
  <div class="figure-caption">
    <strong>Figure 7.</strong> 3D quality space. Drag to rotate. The
    translucent green ceiling marks DR = 100%.
    <a href="fig_07_3d_quality.html">Open standalone →</a>
  </div>
</div>

<h2 id="methodology">11. Methodology</h2>
<div class="notes">
  <ul>
    <li><strong>Detector pool:</strong> TSNFA (proposed cascade); Lipski FFT
    (μ+kσ per-bin energy detector, k = 3 canonical); CA-CFAR (sliding-mean,
    P_fa = 10⁻³); OS-CFAR (75th-percentile rank, P_fa = 10⁻³); CUSUM
    (Tartakovsky variant, α_fa = 10⁻⁵).</li>
    <li><strong>Detection-quality metrics</strong> (DR, precision) are
    <em>per-node algorithmic properties</em> and should be invariant to network
    size.</li>
    <li><strong>Cost metrics</strong> (FP cluster count, network load) scale
    linearly with N because each node operates independently.</li>
    <li><strong>FP clustering window:</strong> consecutive false-positive
    triggers within 5 seconds collapse to a single cluster (operationally
    meaningful FP count).</li>
    <li><strong>ROC curves</strong> are upper-envelope (Pareto frontier) —
    guaranteed monotonic. Hollow circles mark canonical operating points
    (threshold ≈ 1.0).</li>
    <li><strong>Statistical proxies (CDR-100, AST-100):</strong> not included
    in this paper; deferred to a follow-up study with full PyTorch
    implementations.</li>
  </ul>
</div>

<footer>
  Generated by <code>make_data_page.py</code> from
  <code>{workbook_basename}</code> · GNACODE INC, January 2026
</footer>

</div>
</body>
</html>
"""


def build_index_html(runs: List[Dict],
                     figures: List[Tuple[str, str, str, str]],
                     generated_at: str,
                     workbook_basename: str,
                     workbook_path: Path) -> str:
    """Build the rich data-page index from the workbook's pre-computed sheets.

    Tables are extracted from Headline_Metrics, Network_Load_Comparison,
    SNR_Robustness, and Configurations sheets. Figures are embedded as iframes
    so the index page itself stays small (~50 KB) while the figure widgets
    load on demand.
    """
    # Re-open workbook to extract the pre-computed tables.
    wb = load_workbook(workbook_path, data_only=True)

    headline = read_headline_metrics(wb)
    network = read_network_load_comparison(wb)
    snr_data = read_snr_robustness(wb)
    config_rows = read_algorithm_parameters(wb)

    # Render tables
    run_inventory_table = render_run_inventory_table(runs)
    algorithm_params_table = render_algorithm_params_table(config_rows)
    headline_table = render_headline_table(headline, runs)
    network_load_table = render_network_load_table(network)
    snr_robustness_table = render_snr_robustness_table(snr_data)

    config_summary = ' · '.join(sorted({run_label(r) for r in runs}))

    return INDEX_HTML_TEMPLATE.format(
        n_runs=len(runs),
        generated_at=generated_at,
        workbook_basename=workbook_basename,
        config_summary=config_summary,
        run_inventory_table=run_inventory_table,
        algorithm_params_table=algorithm_params_table,
        headline_table=headline_table,
        network_load_table=network_load_table,
        snr_robustness_table=snr_robustness_table,
    )
# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Build interactive HTML data page from compiled_workbook.xlsx',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument('--workbook', default='U:/MONTECARLO/data/compiled_workbook.xlsx',
                        help='Path to compiled_workbook.xlsx (the output of compile_workbook.py)')
    parser.add_argument('--output-dir', default=None,
                        help='Output directory (defaults to <workbook-dir>/datapage)')

    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    output_dir = (Path(args.output_dir).resolve() if args.output_dir
                  else workbook_path.parent / 'datapage')

    print('=' * 70)
    print('TSNFA Monte Carlo — Build Interactive Data Page from Workbook')
    print('=' * 70)
    print(f"  Workbook:   {workbook_path}")
    print(f"  Output dir: {output_dir}")
    print()

    print("Loading workbook...")
    runs = collect_runs(workbook_path)
    if not runs:
        print("\nNo runs found in workbook.")
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)

    figures_meta = [
        ('fig_01_headline_matrix.html', 'Figure 1',
         'Detection Rate Matrix',
         'Heat-map overview: detection rate for each detector across all configurations.'),
        ('fig_02_roc_envelope.html', 'Figure 2',
         'ROC Envelope',
         'Receiver Operating Characteristic curves per configuration. Toggle detectors with the legend.'),
        ('fig_03_network_load.html', 'Figure 3',
         'Network Bandwidth Cost',
         'Per-node and total mesh bandwidth consumption per detector and configuration.'),
        ('fig_04_snr_robustness.html', 'Figure 4',
         'SNR Robustness',
         'Detection rate vs SNR. Reveals OS-CFAR collapse at low SNR; TSNFA invariance.'),
        ('fig_05_scaling.html', 'Figure 5',
         'Network Scaling',
         'How each metric scales from 10n to 50n at each SNR. Quality metrics invariant; cost scales linearly.'),
        ('fig_06_bandwidth_precision.html', 'Figure 6',
         'Bandwidth–Precision Trade-off',
         'Scatter plot revealing TSNFA in the upper-left "precise + cheap" corner.'),
        ('fig_07_3d_quality.html', 'Figure 7',
         '3D Quality Space',
         'FAR × Network Load × Detection Rate. Each detector trajectory connects its four operating points through quality-cost-noise space; drag to rotate.'),
    ]

    builders = [
        fig_01_headline_matrix,
        fig_02_roc_envelope,
        fig_03_network_load,
        fig_04_snr_robustness,
        fig_05_scaling,
        fig_06_bandwidth_precision,
        fig_07_3d_quality,
    ]

    print("\nBuilding figures...")
    for (filename, _, title, _), builder in zip(figures_meta, builders):
        try:
            fig = builder(runs)
            out_path = output_dir / filename
            fig.write_html(
                out_path,
                full_html=True,
                include_plotlyjs='inline',
                config={'displaylogo': False, 'toImageButtonOptions': {'format': 'png', 'scale': 2}},
            )
            print(f"  {filename}  ({out_path.stat().st_size / 1024:.0f} kB)")
        except Exception as e:
            print(f"  {filename}: FAILED ({type(e).__name__}: {e})")
            import traceback
            traceback.print_exc()

    print("\nBuilding index.html...")
    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    index_html = build_index_html(runs, figures_meta, generated_at, workbook_path.name, workbook_path)
    (output_dir / 'index.html').write_text(index_html, encoding='utf-8')
    print(f"  index.html  ({(output_dir / 'index.html').stat().st_size / 1024:.0f} kB)")

    print("\nSaving raw data...")
    serialisable = []
    for r in runs:
        methods_serial = {}
        for method, metrics in r['methods'].items():
            methods_serial[method] = {
                k: {'mean': v[0], 'std': v[1]} for k, v in metrics.items()
            }
        serialisable.append({
            'run_label':        r.get('run_label'),
            'origin':           r.get('origin'),
            'num_nodes':        r.get('num_nodes'),
            'snr_db':           r.get('snr_db'),
            'duration_hours':   r.get('duration_hours'),
            'preset':           r.get('preset'),
            'num_mc_runs':      r.get('num_mc_runs'),
            'total_events':     r.get('total_events'),
            'methods':          methods_serial,
            'roc_sweep':        r.get('roc_sweep', {}),
            'algorithm_params': {
                'tsnfa_gamma_d':  r.get('tsnfa_gamma_d'),
                'tsnfa_gamma_a':  r.get('tsnfa_gamma_a'),
                'tsnfa_zeta':     r.get('tsnfa_zeta'),
                'lipski_k':       r.get('lipski_k'),
                'cacfar_p_fa':    r.get('cacfar_p_fa'),
                'oscfar_k_rank':  r.get('oscfar_k_rank'),
                'oscfar_p_fa':    r.get('oscfar_p_fa'),
                'cusum_alpha_fa': r.get('cusum_alpha_fa'),
            },
        })
    (output_dir / 'data.json').write_text(json.dumps({
        'generated_at':  generated_at,
        'source':        workbook_path.name,
        'num_runs':      len(serialisable),
        'runs':          serialisable,
    }, indent=2, default=str), encoding='utf-8')
    print(f"  data.json  ({(output_dir / 'data.json').stat().st_size / 1024:.0f} kB)")

    print(f"\nData page → {output_dir}/index.html")
    print('Done.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())